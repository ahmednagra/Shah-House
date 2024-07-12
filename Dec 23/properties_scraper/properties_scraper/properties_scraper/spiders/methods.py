import os
from datetime import datetime

from openpyxl.reader.excel import load_workbook  # Load  worksheet
from openpyxl.utils import get_column_letter  # Font Color
from openpyxl.workbook import Workbook  # Worksheet related Tasks

error_list = []


def comparison_excel(data):
    # Define the output directory
    output_dir = 'output/'
    # Create the output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    # Define the file name for the Excel file
    file_name = f'{output_dir}Data.xlsx'

    try:
        # Check if the Excel file exists
        if os.path.isfile(file_name):
            # If the file exists, load the workbook
            wb = load_workbook(file_name)
        else:
            # If the file doesn't exist, create a new workbook
            wb = Workbook()

        # Check if a sheet named "Sheet" exists in the workbook
        if "Sheet" in wb.sheetnames:
            # If the sheet exists, remove it from the workbook
            sheet = wb["Sheet"]
            wb.remove(sheet)

        # Define headers for different types of data
        headers = ['Address', 'street number', 'Type', 'Rooms', 'Other', 'Size (m2)', 'Agency url',
                   'Agency name']
        fulldata_headers = ['Id', 'Address', 'street number', 'Address2', 'street number2', 'Address3',
                            'street number3',
                            'housing company name', 'Zip code', 'Stairway', 'Apartment nr', 'Type', 'Rooms', 'Kitchen',
                            'Bathroom', 'Sauna', 'Other', 'Size (m2)', 'Additional size (m2)', 'Floor', 'Shares',
                            'Nr of shares', 'Agency url', 'Agency name', 'Owned', 'Claimed', 'Status', 'Messages',
                            'Hero images', 'Description', 'building image', ]
        addresses_headers = ['Address', 'street number', 'Address2', 'street number2', 'Address3',
                             'street number3', ]

        # Define the sheets
        fulldata_sheet = get_work_sheet(sheet_name='FULLDATA', wb=wb, headers=fulldata_headers)
        scrapedata_sheet = get_work_sheet(sheet_name='SCRAPEDATA', wb=wb, headers=headers)
        addresses_sheet = get_work_sheet(sheet_name='ADDRESSES', wb=wb, headers=addresses_headers)
        storage_sheet = get_work_sheet(sheet_name='STORAGE', wb=wb, headers=headers + ['timestamp'])

        # Define storage sheet with headers and a timestamp column
        timestamp_format = "%Y-%m-%d %H:%M:%S"

        # ********* step 1  ***************

        # Remove all rows in FULLDATA worksheet where there is any content in column S (agency_url)
        remove_fulldata_agency_url_columns(fulldata_sheet)

        # ********* Step 2: Remove All Rows in SCRAPEDATA ***************

        # Check if 'SCRAPEDATA' sheet exists in the workbook
        # Delete all rows except header row
        if 'SCRAPEDATA' in wb.sheetnames:
            scrapedata_sheet = wb['SCRAPEDATA']
            scrapedata_sheet.delete_rows(2, scrapedata_sheet.max_row)
            print('All Rows Deleted from SCRAPEDATA Sheet Successfully')

        # ********* Step 3: Add Scrape Results from Spiders into SCRAPEDATA worksheet ***************
        for item in data:
            if not item.get('Agency url', ''):
                continue  # Skip this item if 'Agency url' is empty

            street_no = item.get('street number')
            if street_no is not None and street_no.isdigit():
                item['street number'] = int(street_no)

            scrapedata_sheet.append([item.get(col) for col in headers])
        print('All Results added in SCRAPEDATA Sheet Successfully')

        # ********* Step 4: Add Scrape sheet row into FullData sheet with address and street Comparison ***************

        # Check if 'Address' and 'street number' columns are in fulldata_headers
        if "Address" in fulldata_headers and "street number" in fulldata_headers:
            # Column indexes are 1-based
            column_address = fulldata_headers.index("Address") + 1  # Adding 1 because column indexes are 1-based
            column_street_number3 = fulldata_headers.index('street number3') + 1
        else:
            # Use headers from scrapedata if not found in fulldata_headers
            column_address = headers.index("Address") + 1  # Use scrapedata_headers
            column_street_number3 = headers.index("street number3") + 1

        # Initialize a list to store rows from fulldata
        input_fulldata_rows = [{}]

        all_fulldata_addresses = get_sheet_all_addresses_pairs(workbook=wb, sheet_name=fulldata_sheet)
        all_addresses_rows_address = get_sheet_all_addresses_pairs(workbook=wb, sheet_name=addresses_sheet)

        # Iterates row by row from Scrapedata worksheet for further processing.
        for scrapeddata_row in scrapedata_sheet.iter_rows(min_row=2, max_row=scrapedata_sheet.max_row):

            scrapeddata_values_appendable = [cell.value for cell in scrapeddata_row]
            scrapeddata_values = dict(zip(headers, (cell.value for cell in scrapeddata_row)))
            scrapeddata_address = extract_address(scrapeddata_values, 'Address', 'street number')

            print(
                f"Row No {scrapeddata_row[0].row} and {scrapeddata_values_appendable[6]} from scrapedata sheet is called")

            # Initialize variables for matching rows and existing agency URLs
            matching_row_num = None
            fulldata_existing_agency_urls = [row.get('Agency url') for row in input_fulldata_rows]

            # check for fulldata sheet is not empty and columns are within range
            if (fulldata_headers and column_address <= len(fulldata_headers) and column_street_number3
                    <= len(fulldata_headers)):

                unique_scrapped_key = {scrapeddata_values.get('Address', ''): str(
                    scrapeddata_values.get('street number', ''))}

                if any(unique_scrapped_key.items() <= address_row.items() for address_row in all_fulldata_addresses):
                    # Iterate through rows from fulldata_sheet
                    column_agency_name = fulldata_headers.index('Agency name') + 1
                    for fulldata_row in fulldata_sheet.iter_rows(min_row=2, max_row=fulldata_sheet.max_row,
                                                                 min_col=fulldata_sheet.min_column,
                                                                 max_col=column_agency_name):

                        # Create a dictionary with headers as keys and fulldata_values as values
                        fulldata_row_dict = dict(zip(fulldata_headers, (cell.value for cell in fulldata_row)))

                        # Append the dictionary to the list
                        input_fulldata_rows.append(fulldata_row_dict)

                        address1 = extract_address(fulldata_row_dict, 'Address', 'street number')
                        address2 = extract_address(fulldata_row_dict, 'Address2', 'street number2')
                        address3 = extract_address(fulldata_row_dict, 'Address3', 'street number3')

                        # Create list of lowercase addresses
                        fulldata_address = [value.lower() for value in [address1, address2, address3]]

                        # if scrapeddata_address in fulldata_address:

                        if any(scrapeddata_address == address for address in fulldata_address):
                            print(
                                f"ScrapeData Sheet Row_id {scrapeddata_row[0].row} is Matched with FullData Sheet "
                                f"Row_id {fulldata_row[0].row}")
                            matching_row_num = fulldata_row[0].row

                            # Shift existing rows down to make space for the new row
                            fulldata_sheet.insert_rows(matching_row_num, 1)

                            # Create a new row with the existing data
                            matched_fulldata_row = fulldata_row_dict.copy()

                            # Update matched_fulldata_row based on matched address
                            if address1 == scrapeddata_address:
                                matched_fulldata_row['Address'] = scrapeddata_values.get('Address')
                                matched_fulldata_row['street number'] = scrapeddata_values.get('street number')
                            elif address2 == scrapeddata_address:
                                matched_fulldata_row['Address2'] = scrapeddata_values.get('Address')
                                matched_fulldata_row['street number2'] = scrapeddata_values.get('street number')
                            elif address3 == scrapeddata_address:
                                matched_fulldata_row['Address3'] = scrapeddata_values.get('Address')
                                matched_fulldata_row['street number3'] = scrapeddata_values.get('street number')
                            else:
                                matched_fulldata_row['Address'] = scrapeddata_values.get('Address')
                                matched_fulldata_row['street number'] = scrapeddata_values.get('street number')

                            # Update matched_fulldata_row with scrapeddata_values_appendable
                            for header, value in zip(headers, scrapeddata_values_appendable):
                                # Ignore 'Address' and 'street number' because they're already updated
                                if header != 'Address' and header != 'street number':
                                    matched_fulldata_row[header] = value

                            # Add matched_fulldata_row to the fulldata_sheet
                            for header, value in matched_fulldata_row.items():
                                if header in fulldata_headers:
                                    col_index = fulldata_headers.index(header) + 1
                                    if not scrapeddata_values.get('Agency url') in fulldata_existing_agency_urls:
                                        fulldata_sheet.cell(row=matching_row_num, column=col_index, value=value)

                            # Append the matched row's data to the list for future comparison
                            input_fulldata_rows.append(matched_fulldata_row)
                            break
                else:
                    # Check if it's still None (no match in FULL DATA Sheet)
                    if matching_row_num is None:

                        # Initialize a flag to track whether the row is matched in the addresses worksheet
                        matched_in_addresses = False

                        # ********* Step 5: Add row from Scrapedata sheet to Fulldata sheet
                        # if  address and Street no match from Addresses Sheet row  to Scrapedata sheet row *********

                        if any(unique_scrapped_key.items() <= d.items() for d in all_addresses_rows_address):
                            for addresses_row in addresses_sheet.iter_rows(min_row=2, max_row=addresses_sheet.max_row,
                                                                           min_col=addresses_sheet.min_column,
                                                                           max_col=addresses_sheet.max_column):
                                print(f"Row No {addresses_row[0].row} from Addresses sheet is called")
                                row_dict = dict(zip(addresses_headers, (cell.value for cell in addresses_row)))

                                address1 = extract_address(row_dict, 'Address', 'street number')
                                address2 = extract_address(row_dict, 'Address2', 'street number2')
                                address3 = extract_address(row_dict, 'Address3', 'street number3')

                                if scrapeddata_address == address1 or scrapeddata_address == address2 or scrapeddata_address == address3:
                                    data_to_append = []

                                    for header in fulldata_headers:
                                        if header in headers:
                                            col_index = headers.index(header) + 1
                                            if col_index <= len(scrapeddata_values):
                                                value = scrapeddata_values_appendable[col_index - 1]
                                                if value == 'none' or value == '':
                                                    data_to_append.append(None)  # Replace empty string with None

                                                else:
                                                    data_to_append.append(value)
                                            else:
                                                # Append None if no matching value found in scrapeddata_values
                                                data_to_append.append(None)
                                        else:
                                            data_to_append.append(None)  # Append None for headers that don't match

                                    # Append the data to FULLDATA Sheet
                                    if not dict(zip(fulldata_headers, data_to_append)).get('Agency url',
                                                                                           '') in fulldata_existing_agency_urls:
                                        fulldata_sheet.append(data_to_append)
                                        input_fulldata_rows.append(dict(zip(headers, scrapeddata_values_appendable)))
                                        matched_in_addresses = True
                                        break

                        if not matched_in_addresses:
                            print(
                                f'Scrapedata row {scrapeddata_values} not matched in Addresses worksheet and Ignored.')

        # ********* Step 6: Add row in storage sheet from Fulldata Sheet *********

        # Insert rows into Storage those are added in Full data and remove the duplicates
        save_from_fulldata_to_storage_sheet(storage_sheet, fulldata_sheet, fulldata_headers, headers, timestamp_format)

        # set the column widths size values
        set_column_widths(wb=wb)

        wb.save(file_name)
        print(f" Successfully created output File :{file_name}")
        write_errors()
    except Exception as e:
        print(f"An error occurred while saving the data: {str(e)}")
        error = f"An error occurred while saving the data: {str(e)}, FileName : {file_name}"
        error_list.append(error)
        write_errors()

    return file_name


def get_work_sheet(sheet_name, wb, headers):
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(sheet_name)
        sheet.append(headers)

    return sheet


def remove_fulldata_agency_url_columns(fulldata_sheet):
    fulldata_headers = [cell.value for cell in fulldata_sheet[1]]
    claimed_column = fulldata_headers.index('Claimed') + 1
    if fulldata_headers and 'Agency url' in fulldata_headers:
        rows_to_delete = []
        for row in fulldata_sheet.iter_rows(min_row=2, max_row=fulldata_sheet.max_row,
                                            min_col=fulldata_sheet.min_column, max_col=claimed_column):
            current_row_data = {}
            for cell in row:
                column_letter = fulldata_sheet.cell(row=1, column=cell.column).value
                current_row_data[column_letter] = cell.value

            agency_url = current_row_data.get('Agency url', '')
            claimed_colum = current_row_data.get('Claimed', '')
            if claimed_colum is not None:
                claimed_colum = claimed_colum.lower()

            if agency_url and claimed_colum != 'x':
                rows_to_delete.append(row[0].row)

        rows_to_delete.sort(reverse=True)
        for row_id in rows_to_delete:
            fulldata_sheet.delete_rows(idx=row_id)
            print(f"Row No {row_id} is Deleted")

        print(f'Removed {len(rows_to_delete)} rows')

    return print('All Unnecessary Rows are Removed From FullData Sheet')


def get_sheet_all_addresses_pairs(workbook, sheet_name):
    """
    Read all rows from the specified sheet in the workbook and return a list of dictionaries.

    Parameters:
    - workbook: An openpyxl Workbook object.
    - sheet_name: The name of the sheet to read.

    Returns:
    - A list of dictionaries representing each row's data.
    """

    # Initialize an empty list to store row data
    all_rows = []

    # Iterate over rows in the sheet
    for row in sheet_name.iter_rows(values_only=True):
        # Construct a dictionary representing the row's data
        row_data = {}
        for index, value in enumerate(row):
            # Get the column letter corresponding to the column index
            column_letter = get_column_letter(index + 1)
            # Get the header from the first row of the sheet
            header = sheet_name[f"{column_letter}1"].value
            # Add the header and its corresponding value to the row's data dictionary
            row_data[header] = value
        # Append the row's data dictionary to the list
        all_rows.append(row_data)

    all_full_data_address = []
    for row in all_rows:
        # Construct unique keys for fulldata addresses
        address_street_pairs = {
            row.get('Address', ''): str(row.get('street number', '')),
            row.get('Address2', ''): str(row.get('street number2', '')),
            row.get('Address3', ''): str(row.get('street number3', ''))
        }

        all_full_data_address.append(address_street_pairs)

    return all_full_data_address


def extract_address(row, *keys):
    try:
        address = '_'.join(str(row.get(key, '')).lower() if row.get(key) is not None else '' for key in keys)
        return address
    except Exception as e:
        error = f'function extract_address  arises this error :{e}'
        error_list.append(error)
        address = ''
        return address


def set_column_widths(wb):
    """
    Set column widths for sheets in the workbook.

    :param wb: Workbook object
    """
    try:
        column_widths = {
            'Address': 25,
            'street number': 15,
            'Address2': 25,
            'street number2': 15,
            'Address3': 25,
            'street number3': 15,
            'Type': 20,
            'Rooms': 10,
            'Other': 10,
            'Size (m2)': 15,
            'Agency url': 50,
            'Agency name': 15,
            'timestamp': 20,
        }

        set_widths = {}  # Initialize a dictionary to store set column widths
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for header_name, width in column_widths.items():
                for column_letter, cell in enumerate(sheet[1], start=1):
                    if cell.value == header_name:
                        if width is not None:
                            sheet.column_dimensions[get_column_letter(column_letter)].width = width
                            set_widths[header_name] = width  # Store the set column width in the dictionary
        return set_widths
    except Exception as e:
        error = f'function set_column_widths  arises this error :{e}'
        error_list.append(error)
        set_widths = {}
        return set_widths


def save_from_fulldata_to_storage_sheet(storage_sheet, fulldata_sheet, fulldata_headers, headers, timestamp_format):
    try:
        # Insert rows into Storage those are added in Full data and remove the duplicates
        storage_existing_data = {}
        for storage_row in storage_sheet.iter_rows(min_row=2, max_row=storage_sheet.max_row, min_col=1,
                                                   max_col=8):  # column_address ==1, Agency name ==7
            key = tuple([cell.value for cell in storage_row])
            storage_existing_data[key] = [cell.value for cell in storage_row]

        # Create a mapping between column names in fulldata_headers and headers
        column_mapping = {}
        for header in fulldata_headers:
            if header in headers:
                column_mapping[header] = headers.index(header)

        # Iterate through rows in FULLDATA
        for row in fulldata_sheet.iter_rows(min_row=2, max_row=fulldata_sheet.max_row,
                                            min_col=1, max_col=len(fulldata_headers)):

            # Skip those rows from FullData , it's already existing
            fulldata_row_dict = dict(zip(fulldata_headers, (cell.value for cell in row)))
            if not fulldata_row_dict.get('Agency url'):
                continue

            fulldata_head = [fulldata_headers[cell.column - 1] for cell in row]
            fulldata_values = [cell.value for cell, header in zip(row, fulldata_head) if header in column_mapping]

            # Check if the row already exists in STORAGE
            if tuple(fulldata_values) not in storage_existing_data:
                timestamp = datetime.now().strftime(timestamp_format)
                storage_sheet.append(fulldata_values + [timestamp])
                storage_existing_data[tuple(fulldata_values)] = fulldata_values

        print("Rows added to Storage Sheet successfully.")
        return storage_existing_data
    except Exception as e:
        error = f'function save_from_fulldata_to_storage_sheet  arises this error :{e}'
        error_list.append(error)


def write_errors():
    with open('METHODS_ERRORS.txt', 'w') as f:
        for error_message in error_list:
            f.write(f"{error_message}\n")
