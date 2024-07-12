import csv
import logging
import os
import glob
import json
from math import ceil
from datetime import datetime
from collections import OrderedDict

from scrapy import Spider, Request

from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook


class UsedcarsSpider(Spider):
    name = "cars"
    start_urls = ['https://www.cargurus.com/']

    logging.basicConfig(filename='example.log', encoding='utf-8', level=logging.INFO)

    headers = ['Title', 'Make', 'Model', 'Year', 'Vin No',
               'Location', 'Condition', 'Price',
               'Average Market Price', 'Main Image', 'Options',
               'Mileage', 'Trim', 'Color', 'Body Type', 'Fuel Type', 'Engine',
               'Deal Rating', 'Days on Market', 'Drive Train', 'City Fuel Economy',
               'Highway Fuel Economy', 'Combined Fuel Economy', 'Dealer Name',
               'Dealer City', 'Dealer State', 'Dealer Zip', 'Dealer Rating',
               # 'Dealer Reviews Count',
               'Dealer Phone', 'URL']

    custom_settings = {
        'AUTOTHROTTLE_ENABLED': True,
        'AUTOTHROTTLE_START_DELAY': 1.0,  # Initial delay in seconds
        'AUTOTHROTTLE_TARGET_CONCURRENCY': 0.5,  # Adjust as needed
        'AUTOTHROTTLE_DEBUG': True,  # Set to True for debugging

        'CONCURRENT_REQUESTS': 3,
        'FEED_EXPORTERS': {'xlsx': 'scrapy_xlsx.XlsxItemExporter'},
        'FEEDS': {
            f'output/Cargurus Used Cars Details.xlsx': {
                'format': 'xlsx',
                'fields': headers,
            }
        },
    }

    def __init__(self, **kwargs):

        super().__init__(**kwargs)
        self.files_rename()
        self.current_scraped_items = []

        self.previously_scraped_items = {item.get('Vin No'): item for item in
                                         self.read_excel_file('output/Previous Cargurus Used Cars Details.xlsx')}

    def parse(self, response, **kwargs):
        # Get the Maker Values and make Request
        companies = response.css('#carPickerUsed_makerSelect optgroup[label="All Makes"] option')

        for company in companies[:1]:
            value = company.css('option::attr(value)').get('')
            name = company.css('option::text').get('').strip()
            formatted_name = '-'.join(name.split())
            url = f'https://www.cargurus.com/Cars/l-Used-{formatted_name}-{value}'

            yield Request(url=url, callback=self.parse_company_pagination, meta={'company': value})

    def parse_company_pagination(self, response):
        url = response.url
        script = response.css('script:contains(totalListings)').re_first(r'({.*})')
        data = json.loads(script)
        total_results = data.get('totalListings', 0)
        print('Total Cars Found :', total_results)

        if total_results > 1:
            total_pages = ceil(total_results / 48)
            print('Total Pages are :', total_pages)

            for page_number in range(1, total_pages + 1):
                offset = page_number * 48
                company = response.meta.get('company', '')

                # Get the Company value from response and make json request for records with pagination
                page_url = f'https://www.cargurus.com/Cars/searchResults.action?entitySelectingHelper.selectedEntity={company}&sourceContext=untrackedExternal_false_0&inventorySearchWidgetType=AUTO&sortDir=ASC&sortType=DEAL_RATING_RPL&shopByTypes=MIX&srpVariation=DEFAULT_SEARCH&nonShippableBaseline=285&offset={offset}&maxResults=48&filtersModified=true'
                yield Request(url=page_url, callback=self.parse_detail_usedcar, dont_filter=True, meta={'url': url})
        else:
            return

    def parse_detail_usedcar(self, response):
        # get response and load into json then get records one by one
        try:
            data = json.loads(response.body)

            if not data or data == None:
                return

            for car in data:
                item = OrderedDict()
                make = car.get('makeName', '')
                year = car.get('carYear', '')
                item['Title'] = car.get('listingTitle', '')
                item['Make'] = car.get('makeName', '')
                item['Model'] = car.get('modelName', '').replace(str(make), '').replace(str(year), '').strip()
                item['Year'] = car.get('carYear', '')
                item['Vin No'] = car.get('vin', '')

                item['Location'] = car.get('sellerCity', '')
                item['Condition'] = 'Used'
                item['Price'] = car.get('priceString', '')
                item['Average Market Price'] = f"${car.get('expectedPrice', '')}"
                item['Main Image'] = car.get('originalPictureData', {}).get('url', '')
                item['Options'] = ', '.join([x for x in car.get('options', '')])

                item['Mileage'] = f"{car.get('mileageString', '')} Miles"
                item['Trim'] = car.get('trimName', '')

                item['Color'] = car.get('exteriorColorName', '') or car.get('normalizedExteriorColor', '')
                item['Body Type'] = car.get('bodyTypeName', '')
                item['Fuel Type'] = car.get('localizedFuelType', '')
                item['Engine'] = car.get('localizedEngineDisplayName', '')

                item['Deal Rating'] = car.get('dealRating', '')
                item['Days on Market'] = car.get('daysOnMarket', '')

                item['Drive Train'] = car.get('driveTrain', '')
                item['City Fuel Economy'] = ' '.join(
                    [str(value) for value in list(car.get('cityFuelEconomy', {}).values())])
                item['Highway Fuel Economy'] = ' '.join(
                    [str(value) for value in list(car.get('highwayFuelEconomy', {}).values())])
                item['Combined Fuel Economy'] = ' '.join(
                    [str(value) for value in list(car.get('combinedFuelEconomy', {}).values())])

                item['Dealer Name'] = car.get('serviceProviderName', '')
                item['Dealer City'] = car.get('sellerCity', '').split(',')[0].strip()
                item['Dealer State'] = car.get('sellerRegion', '')
                item['Dealer Zip'] = car.get('sellerPostalCode', '')
                item['Dealer Rating'] = round(car.get('sellerRating', 0), 2)
                # item['Dealer Reviews Count'] = car.get('reviewCount', 0)
                item['Dealer Phone'] = car.get('phoneNumberString', '')
                item['URL'] = response.meta.get('url') + f"#listing={str(car.get('id', ''))}/NONE/DEFAULT"

                self.current_scraped_items.append(item)
                yield item

        except Exception as e:
            self.logger.error(f"Error while parsing used cars: {e}")

    def close(spider, reason):
        spider.comparison_data()

    def get_work_book(self, filepath, file_name, sheet_name, headers, data):
        try:
            # Create or load the workbook
            if os.path.isfile(filepath):
                wb = load_workbook(filepath)
            else:
                wb = Workbook()

            # Remove default sheet if it exists
            if "Sheet" in wb.sheetnames:
                default_sheet = wb["Sheet"]
                wb.remove(default_sheet)

            # Get the sheet or create a new one
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                sheet = wb.create_sheet(sheet_name)
                # Append headers if the sheet is newly created
                if headers:
                    sheet.append(headers)

            # Write data to the provided sheet
            for row_data in data:
                sheet.append(row_data)

            # Save the workbook
            wb.save(filepath)
            print(f'Successfully wrote the file: {file_name}')

            # Return the workbook or sheet object if needed
            return wb

        except Exception as e:
            print(f"Error occurred while writing the file: {file_name}")
            print(f"Error details: {str(e)}")
            return None

    def comparison_data(self):
        if not self.current_scraped_items:
            return

        current_time = datetime.now()
        formatted_time = current_time.strftime("%Y-%m-%d_%H-%M-%S-%f")[:-2]

        # Specify the file name and output folder
        output_folder = 'output/Reports'

        # Create the output folder if it doesn't exist
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        sold_file_name = f'Sold_Cars_{formatted_time}.xlsx'
        master_file_name = f'Master SoldCars.xlsx'
        sold_cars_filepath = os.path.join(output_folder, sold_file_name)
        master_sold_cars_filepath = os.path.join(output_folder, master_file_name)

        current_scraped_vin = [item.get('Vin No') for item in self.current_scraped_items]
        sold_cars = []

        # Iterate through previous items, and append to sold_car sheet if VIN not in current_scraped_vin
        for previous_vin, previous_item in self.previously_scraped_items.items():
            if previous_vin not in current_scraped_vin:
                row_values = list(previous_item.values())
                sold_cars.append(row_values)

        # Write data to the sold cars workbook
        sold_cars_written = self.get_work_book(sold_cars_filepath, sold_file_name, sheet_name='Sold Cars',
                                               headers=self.headers, data=sold_cars)

        if sold_cars:
            # Append data to the master sold cars workbook if sold_cars is not empty
            master_sold_cars_written = self.get_work_book(master_sold_cars_filepath, master_file_name,
                                                          sheet_name='Master Sold Cars', headers=self.headers,
                                                          data=sold_cars)
        else:
            master_sold_cars_written = False

        if sold_cars_written and master_sold_cars_written:
            print('Files Written Successfully')
            return True
        else:
            print('Error writing files')
            return False

    def files_rename(self):
        all_files = glob.glob('output/*')
        # Check if any file has "previous" in the filename and delete it
        for file_path in all_files:
            if "previous" in file_path.lower():
                os.remove(file_path)

        # Find the remaining file
        remaining_files = glob.glob('output/*')
        for file in remaining_files:
            if 'Previous' not in file and 'Cargurus Used Cars Details' in file:

                # Rename the remaining file by adding "Previous" to the filename
                new_file_path = os.path.join('output', f'Previous {os.path.basename(file)}')
                os.rename(file, new_file_path)
                return

            else:
                return

    def read_excel_file(self, file_path):
        data = []
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                data.append(row_data)

        except Exception as e:
            print(f"Error reading Excel file: {e}")

        return data
