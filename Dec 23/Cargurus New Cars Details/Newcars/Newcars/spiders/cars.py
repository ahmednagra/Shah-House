import logging
import os
import glob
from math import ceil
from datetime import datetime
from collections import OrderedDict

from scrapy import Spider, Request

from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook


class NewcarsSpider(Spider):
    name = "cars"
    start_urls = [
        'https://www.cargurus.com/Cars/getCarPickerReferenceDataAJAX.action?localCountryCarsOnly=true&outputFormat=REACT&showInactive=false&useInventoryService=true&quotableCarsOnly=true']

    # logging.basicConfig(filename='example.log', encoding='utf-8', level=logging.INFO)

    headers = ['Title', 'Make', 'Model', 'Year', 'Vin No',
               'Location', 'Condition', 'Price',
               'Average Market Price', 'Main Image', 'Options',
               'Mileage', 'Trim', 'Color', 'Body Type', 'Fuel Type', 'Engine',
               'Deal Rating', 'Days on Market', 'Drive Train', 'City Fuel Economy',
               'Highway Fuel Economy', 'Combined Fuel Economy', 'Dealer Name',
               'Dealer City', 'Dealer State', 'Dealer Zip', 'Dealer Rating',
               # 'Dealer Reviews Count',
               'Dealer Phone', 'URL']

    custom_settings = {
        # 'CONCURRENT_REQUESTS': 4,
        'FEED_EXPORTERS': {'xlsx': 'scrapy_xlsx.XlsxItemExporter'},
        'FEEDS': {
            f'output/Cargurus New Cars Details.xlsx': {
                'format': 'xlsx',
                'fields': headers,
            }
        },
    }

    def __init__(self, **kwargs):

        super().__init__(**kwargs)
        self.files_rename()
        self.current_scraped_items = []

        self.previously_scraped_items = {item.get('Vin No'): item for item in
                                         self.read_excel_file('output/Previous Cargurus New Cars Details.xlsx')}

        self.zipcodes = self.read_uszip_file()

    def parse(self, response, **kwargs):

        data = response.json().get('allMakerModels', []).get('makers', [{}])

        # for zipcode in self.zipcodes[:1]:
        #     zipcode = '85001'
        #     url = f'https://www.cargurus.com/Cars/preflightResults.action?searchId=842cfc71-84f0-44a1-9ca9-0f565e9a95ec&zip={zipcode}&distance=50&entitySelectingHelper.selectedEntity=&sourceContext=untrackedExternal_false_0&isNewCarSearch=true&newUsed=1&inventorySearchWidgetType=NEW_CAR&sortDir=ASC&sortType=PRICE&shopByTypes=MIX&srpVariation=NEW_CAR_SEARCH&nonShippableBaseline=0'
        #
        #     # for maker in [x.get('id') for x in data]:
        #     #     print(maker)
        #         # url = f'https://www.cargurus.com/Cars/preflightResults.action?searchId=842cfc71-84f0-44a1-9ca9-0f565e9a95ec&zip={zipcode}&distance=50&entitySelectingHelper.selectedEntity={maker}&sourceContext=untrackedExternal_false_0&isNewCarSearch=true&newUsed=1&inventorySearchWidgetType=NEW_CAR&sortDir=ASC&sortType=PRICE&shopByTypes=MIX&srpVariation=NEW_CAR_SEARCH&nonShippableBaseline=0'
        #     yield Request(url=url, callback=self.pagination)

        for zipcode in self.zipcodes[:1]:
            zipcode = '85001'
            for maker in [x.get('id') for x in data]:
                url = f'https://www.cargurus.com/Cars/preflightResults.action?searchId=842cfc71-84f0-44a1-9ca9-0f565e9a95ec&zip={zipcode}&distance=50&entitySelectingHelper.selectedEntity={maker}&sourceContext=untrackedExternal_false_0&isNewCarSearch=true&newUsed=1&inventorySearchWidgetType=NEW_CAR&sortDir=ASC&sortType=PRICE&shopByTypes=MIX&srpVariation=NEW_CAR_SEARCH&nonShippableBaseline=0'
                yield Request(url=url, callback=self.pagination)

    def pagination(self, response):
        res = response.json()
        total_results = res.get('totalListings', 0)
        print('Total Cars Found :', total_results)

        if total_results == 0:
            return

        if total_results > 0:
            total_pages = ceil(total_results / 15)
            print('Total Pages are :', total_pages)

            for page_number in range(total_pages):
                offset = page_number * 15
                maker = res.get('search', {}).get('selectedEntity', '')
                zip_no = res.get('search', {}).get('zip', '')
                # page_url =f'https://www.cargurus.com/Cars/preflightResults.action?searchId=b16cd2aa-db81-45cd-a700-617cbe7f6e36&zip={zip_no}&distance=50&entitySelectingHelper.selectedEntity={maker}&sourceContext=untrackedWithinSite_false_0&isNewCarSearch=true&newUsed=1&inventorySearchWidgetType=NEW_CAR&sortDir=ASC&sortType=PRICE&shopByTypes=MIX&srpVariation=NEW_CAR_SEARCH&nonShippableBaseline={total_results}&offset={offset}&maxResults=15&filtersModified=true'
                page_url =f'https://www.cargurus.com/Cars/preflightResults.action?searchId=b16cd2aa-db81-45cd-a700-617cbe7f6e36&zip={zip_no}&distance=50&entitySelectingHelper.selectedEntity={maker}&sourceContext=untrackedWithinSite_false_0&isNewCarSearch=true&newUsed=1&inventorySearchWidgetType=NEW_CAR&sortDir=ASC&sortType=PRICE&shopByTypes=MIX&srpVariation=NEW_CAR_SEARCH&nonShippableBaseline={total_results}&offset={offset}&maxResults=15&filtersModified=true'

                print(page_url)
                # yield Request(url=page_url, callback=self.parse_newcars, dont_filter=True)
                yield Request(url=page_url, callback=self.parse_newcars)
        else:
            return

    def parse_newcars(self, response):
        if response.text.strip() == 'null':
            print('No reulst founded in url :', response.url)
            return

        # get response and load into json then get records one by one
        try:
            data = response.json()
            desired_keys = ['listings', 'priorityListings', 'highlightListings', 'newCarFeaturedListings']

            combine_dict = {key: data.get(key, {}) for key in desired_keys}
            new_list = []
            [new_list.extend(value) for value in combine_dict.values()]
            for car in new_list:
                item = OrderedDict()
                car_id = car.get('id', 0)

                title = car.get('listingTitle', '')
                if not title:
                    continue

                vin_number = car.get('vin', '')
                if any(item.get('Vin No', '') == vin_number for item in self.current_scraped_items):
                    print(vin_number, 'product already exists')
                    continue

                item['Title'] = car.get('listingTitle', '')
                item['Make'] = car.get('makeName', '')
                item['Model'] = car.get('modelName', '')
                item['Year'] = car.get('carYear', 0)
                item['Vin No'] = car.get('vin', '')

                item['Location'] = car.get('sellerCity', '')
                item['Condition'] = 'New'
                item['Price'] = car.get('priceString', '')
                # item['Average Market Price'] = f"${car.get('msrp', 0)}"
                item['Average Market Price'] = f"${car['msrp']}" if car.get('msrp') is not None else ''
                item['Main Image'] = car.get('originalPictureData', {}).get('url', '')
                item['Options'] = ', '.join([x for x in car.get('options', '')])

                item['Mileage'] = f"{car.get('mileageString', '')} Miles"
                item['Trim'] = car.get('trimName', '')

                item['Color'] = car.get('normalizedExteriorColor', '')
                item['Body Type'] = car.get('bodyTypeName', '')
                item['Fuel Type'] = car.get('localizedFuelType', '')
                item['Engine'] = car.get('localizedEngineDisplayName', '')

                item['Deal Rating'] = car.get('dealRating', '')
                item['Days on Market'] = car.get('daysOnMarket', '')

                item['Drive Train'] = car.get('driveTrain', '')
                item['City Fuel Economy'] = ' '.join(
                    [str(value) for value in list(car.get('cityFuelEconomy', {}).values())])
                item['Highway Fuel Economy'] = ' '.join(
                    [str(value) for value in list(car.get('highwayFuelEconomy', {}).values())])
                item['Combined Fuel Economy'] = ' '.join(
                    [str(value) for value in list(car.get('combinedFuelEconomy', {}).values())])

                item['Dealer Name'] = car.get('serviceProviderName', '')
                item['Dealer City'] = car.get('sellerCity', '').split(',')[0].strip()
                item['Dealer State'] = car.get('sellerRegion', '')
                item['Dealer Zip'] = car.get('sellerPostalCode', '')
                item['Dealer Rating'] = round(car.get('sellerRating', 0), 2)
                item['Dealer Phone'] = car.get('phoneNumberString', '')
                item[
                    'URL'] = f'https://www.cargurus.com/Cars/inventorylisting/vdp.action?listingId={car_id}&entitySelectingHelper.selectedEntity=NEW_CAR#listing={car_id}/NEWCAR_FEATURED/DEFAULT'

                self.current_scraped_items.append(item)
                yield item

        except Exception as e:
            self.logger.error(f"Error while parsing used cars: {e}")

    def close(spider, reason):
        spider.comparison_data()

    def get_work_sheet(self, sheet_name, wb, headers):

        # function for worksheet select or name the worksheet
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)
            sheet.append(headers)

        return sheet

    def comparison_data(self):
        if not self.current_scraped_items:
            return

        current_time = datetime.now()
        formatted_time = current_time.strftime("%Y-%m-%d_%H-%M-%S-%f")[:-2]

        # Specify the file name and output folder
        output_folder = 'output/Reports'
        file_name = f'New Cars_{formatted_time}.xlsx'
        file_path = os.path.join(output_folder, file_name)

        # Create the output folder if it doesn't exist
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Create or load the workbook
        if os.path.isfile(file_path):
            wb = load_workbook(file_path)
        else:
            wb = Workbook()

        if "Sheet" in wb.sheetnames:
            sheet = wb["Sheet"]
            wb.remove(sheet)

        # # Create or get sheets
        new_cars = self.get_work_sheet(sheet_name='New Cars Record', wb=wb, headers=self.headers)

        # Get the list of VINs from current_scraped_items
        current_scraped_vin = [item.get('Vin No') for item in self.current_scraped_items]

        # Iterate through previous items, and append to sold_car sheet if VIN not in current_scraped_vin
        for previous_vin, previous_item in self.previously_scraped_items.items():
            if previous_vin not in current_scraped_vin:
                row_values = list(previous_item.values())
                new_cars.append(row_values)

        # Save the workbook
        wb.save(file_path)

    def files_rename(self):
        all_files = glob.glob('output/*')
        # Check if any file has "previous" in the filename and delete it
        for file_path in all_files:
            if "previous" in file_path.lower():
                os.remove(file_path)

        # Find the remaining file
        remaining_files = glob.glob('output/*')
        for file in remaining_files:
            if 'Previous' not in file and 'Cargurus New Cars Details' in file:

                # Rename the remaining file by adding "Previous" to the filename
                new_file_path = os.path.join('output', f'Previous {os.path.basename(file)}')
                os.rename(file, new_file_path)
                return

            else:
                return

    def read_excel_file(self, file_path):
        data = []
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                data.append(row_data)

        except Exception as e:
            print(f"Error reading Excel file: {e}")

        return data

    def read_uszip_file(self):
        file_path = ''.join(glob.glob('input/uszip.txt'))
        rows = []
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            for line in file:
                rows.append(line.strip())
        return rows
