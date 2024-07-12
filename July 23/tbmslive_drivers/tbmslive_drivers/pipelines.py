# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter

from datetime import datetime, timedelta
from openpyxl.styles import PatternFill
from scrapy.exceptions import DropItem
from scrapy_xlsx.exporters import XlsxItemExporter
from openpyxl.styles import colors
from openpyxl import Workbook


RED_COLOR = colors.Color(rgb="FF0000")
YELLOW_COLOR = colors.Color(rgb="FFFF00")
GREEN_COLOR = colors.Color(rgb="00FF00")


# class CustomXlsxItemExporter(XlsxItemExporter):
#
#     def __init__(self, file, **kwargs):
#         self._configure(kwargs)
#         self._file = file
#         self._initialize_workbook()
#         self._initialize_sheet()
#         self.items = []
#
#     def _initialize_workbook(self):
#         self._workbook = Workbook()
#
#     def _initialize_sheet(self):
#         self._sheet = self._workbook.active
#
#     def export_items(self):
#         self._initialize_sheet()
#
#         for item in self.items:
#             row = next(self._sheet.iter_rows())
#             self._write_row(self._sheet, row, item)
#
#         self._workbook.save(self._file)
#
#     def _write_row(self, sheet, row, item):
#         row_data = self._convert_row(item)
#
#         # Customize based on expiration dates
#         expiry_date_format = '%Y-%m-%d'  # Adjust the format based on your date format
#         vehicle_expiry = datetime.strptime(item['vehicle_licence_expiry'], expiry_date_format).date()
#         driver_expiry = datetime.strptime(item['driver_licence_expiry'], expiry_date_format).date()
#
#         today = datetime.now().date()
#         thirty_days_later = today + timedelta(days=30)
#
#         for cell, value in zip(row, row_data):
#             cell.value = value
#
#             # Apply color coding
#             if cell.column == self._headers.index('vehicle_licence_expiry') + 1:
#                 if vehicle_expiry < today:
#                     cell.fill = PatternFill(start_color=RED_COLOR, end_color=RED_COLOR, fill_type='solid')
#                 elif vehicle_expiry < thirty_days_later:
#                     cell.fill = PatternFill(start_color=YELLOW_COLOR, end_color=YELLOW_COLOR, fill_type='solid')
#                 else:
#                     cell.fill = PatternFill(start_color=GREEN_COLOR, end_color=GREEN_COLOR, fill_type='solid')
#
#             elif cell.column == self._headers.index('driver_licence_expiry') + 1:
#                 if driver_expiry < today:
#                     cell.fill = PatternFill(start_color=RED_COLOR, end_color=RED_COLOR, fill_type='solid')
#                 elif driver_expiry < thirty_days_later:
#                     cell.fill = PatternFill(start_color=YELLOW_COLOR, end_color=YELLOW_COLOR, fill_type='solid')
#                 else:
#                     cell.fill = PatternFill(start_color=GREEN_COLOR, end_color=GREEN_COLOR, fill_type='solid')
#

class TbmsliveDriversPipeline:
    def process_item(self, item, spider):
        # Apply item processing logic here

        # Check if the item meets certain conditions
        if item['driver_licence_expiry'] is not None and item['vehicle_licence_expiry'] is not None:
            driver_expiry = datetime.strptime(item['driver_licence_expiry'], '%Y-%m-%d').date()
            vehicle_expiry = datetime.strptime(item['vehicle_licence_expiry'], '%Y-%m-%d').date()
            today = datetime.now().date()
            thirty_days_later = today + timedelta(days=30)

            if driver_expiry < today or vehicle_expiry < today:
                # Item expired, discard it
                raise DropItem("Item expired: %s" % item)
            elif driver_expiry < thirty_days_later or vehicle_expiry < thirty_days_later:
                # Item will expire within 30 days, modify the item as desired
                item['status'] = 'Expiring soon'
            else:
                # Return the processed item
                return item
        else:
            # Item is missing necessary information, discard it
            raise DropItem("Item missing necessary information: %s" % item)




