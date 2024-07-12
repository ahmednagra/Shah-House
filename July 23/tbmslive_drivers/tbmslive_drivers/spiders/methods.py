import os
from datetime import datetime, timedelta

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

RED_COLOR = 'FFFF0000'
YELLOW_COLOR = 'FFFFFF00'
GREEN_COLOR = 'FF00FF00'
AMBER_COLOR = 'FFBF00'

today_date = datetime.now().date()
thirty_days_later = today_date + timedelta(days=30)
expiry_date_format = '%d/%m/%Y'


def write_to_excel(data, sheet_name):
    output_dir = 'output/'
    os.makedirs(output_dir, exist_ok=True)
    file_name = f'{output_dir}Driver License Report {datetime.now().strftime("%d%m%Y%H%M%S")}.xlsx'

    try:
        if os.path.isfile(file_name):
            wb = load_workbook(file_name)
        else:
            wb = Workbook()

        if "Sheet" in wb.sheetnames:
            sheet = wb["Sheet"]
            wb.remove(sheet)

        sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

        headers = ['Name', 'Driver PCO License', 'Licence Holder Name', 'Driver PCO Licence Expiry', 'Veh PCO Licence',
                   'Veh PCO Licence Expiry', 'Vehicle Registration Number (VRM)', 'Vehicle Make', 'Vehicle Model',
                   'Vehicle Plate/ Disc Number']

        sheet.append(list(headers))

        for item in data.values():
            # values = list(item.values())
            sheet.append([item.get(col) for col in headers])

        # Set column widths
        column_widths = [40, 15, 15, 30, 20, 20, 20, 15, 20, 20]

        for i, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = width

        for row in sheet.iter_rows(min_row=2):
            driver_expiry_cell = row[list(headers).index('Driver PCO Licence Expiry')]
            vehicle_expiry_cell = row[list(headers).index('Veh PCO Licence Expiry')]

            driver_expiry_value = driver_expiry_cell.value
            vehicle_expiry_value = vehicle_expiry_cell.value

            driver_expiry_cell_color = get_expiry_cell_color(driver_expiry_value)
            driver_expiry_cell.fill = PatternFill(start_color=driver_expiry_cell_color,
                                                  end_color=driver_expiry_cell_color,
                                                  fill_type='solid')

            vehicle_expiry_cell_color = get_expiry_cell_color(vehicle_expiry_value)
            vehicle_expiry_cell.fill = PatternFill(start_color=vehicle_expiry_cell_color,
                                                   end_color=vehicle_expiry_cell_color,
                                                   fill_type='solid')

        wb.save(file_name)
        print(f"Data saved to {file_name}")

    except Exception as e:
        print(f"An error occurred while saving the data: {str(e)}")

    return file_name


def get_expiry_cell_color(license_expiry_text):
    if license_expiry_text and 'expire' in license_expiry_text.lower():
        return RED_COLOR

    try:
        license_expiry_date = datetime.strptime(license_expiry_text, expiry_date_format).date()
    except Exception as e:
        return AMBER_COLOR

    if license_expiry_date:
        if license_expiry_date <= today_date:
            cell_color = RED_COLOR
        elif license_expiry_date <= thirty_days_later:
            cell_color = YELLOW_COLOR
        else:
            cell_color = GREEN_COLOR
    else:
        cell_color = AMBER_COLOR

    return cell_color


def upload_file_to_drive(file_name):
    service_account_key_path = 'input/credential.json'
    # folder_id = '1bKChTxcp4CIJaZeXllGmLuodNTYH0AeX'
    folder_id = '1k5NFGwK8lK_n07EHdk6IbhdIf78OUiX_'

    credentials = service_account.Credentials.from_service_account_file(
        service_account_key_path,
        scopes=['https://www.googleapis.com/auth/drive.file']
    )
    drive_service = build('drive', 'v3', credentials=credentials)

    file_name = os.path.basename(file_name)
    file_path = 'output/' + file_name
    metadata = {
        'name': file_path,
        'parents': [folder_id]
    }
    media = MediaFileUpload(file_path, resumable=True)
    file = drive_service.files().create(
        body=metadata,
        media_body=media,
        fields='id'
    ).execute()

    print('File uploaded successfully. File ID:', file.get('id'))


