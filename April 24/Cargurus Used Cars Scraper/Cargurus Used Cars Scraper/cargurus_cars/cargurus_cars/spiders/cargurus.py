import csv
import os
import json
from math import ceil
from datetime import datetime
from collections import OrderedDict
import threading

from scrapy import Spider, Request

from openpyxl.reader.excel import load_workbook

from slugify import slugify


class UsedcarsSpider(Spider):
    name = "cars"
    start_urls = ['https://www.cargurus.com/']

    headers = {
        'authority': 'www.cargurus.com',
        'accept': '*/*',
        'accept-language': 'en-US,en;q=0.9',
        'referer': 'https://www.cargurus.com/Cars/l-Used-Acura-m4',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    }

    csv_headers = ['Title', 'Make', 'Model', 'Year', 'Vin',
                   'Location', 'Condition', 'Price',
                   'Average Market Price', 'Main Image', 'Options',
                   'Mileage', 'Trim', 'Color', 'Body Type', 'Fuel Type', 'Engine',
                   'Deal Rating', 'Days on Market', 'Drive Train', 'City Fuel Economy',
                   'Highway Fuel Economy', 'Combined Fuel Economy', 'Dealer Name',
                   'Dealer City', 'Dealer State', 'Dealer Zip', 'Dealer Phone',  'Dealer Rating',
                   'URL', 'Scraped Date']

    main_start_datetime_str = datetime.now().strftime("%Y-%m-%d %H%M%S")

    custom_settings = {
        'CONCURRENT_REQUESTS': 5,
        'DOWNLOAD_DELAY': 1,
        'AUTOTHROTTLE_ENABLED': True,
        'AUTOTHROTTLE_START_DELAY': 1,
        'AUTOTHROTTLE_MAX_DELAY': 10,

        'FEEDS': {
            f'output/CarGurus Used Cars {main_start_datetime_str}.csv': {
                'format': 'csv',
                'fields': csv_headers,
            }
        },
    }

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.scraped_product_counter = 0
        self.error_in_filter = []
        self.csv_lock = threading.Lock()
        self.mandatory_logs = [f'Spider "{self.name}" Started at "{self.main_start_datetime_str}"\n']
        self.logs_filename = f'logs/logs {self.main_start_datetime_str}.txt'
        self.scraper_started_at = datetime.now()

        self.current_scraped_items = []
        self.scraping_datetime = datetime.now().strftime('%Y-%m-%d %H%M%S')

        self.previously_scraped_items = {item.get('Vin'): item for item in self.get_previous_scarped_cars_from_file()}

        a=0

    def parse(self, response, **kwargs):
        for make in response.css('#carPickerUsed_makerSelect optgroup option')[:1]:
            make_id = make.css('::attr(value)').get('')
            make_name = make.css('::text').get('')

            # url = f'https://www.cargurus.com/Cars/l-Used-{slugify(make_name).title()}-{make_id}'

            #  for test purpose
            url = 'https://www.cargurus.com/Cars/l-Used-Ford-m2'

            yield Request(url=url, callback=self.pagination, meta={'company': make_id})

    def pagination(self, response):
        url = response.url
        script = response.css('script:contains(totalListings)').re_first(r'({.*})')
        data = json.loads(script)
        total_results = data.get('totalListings', 0)

        if not total_results:
            return

        total_pages = ceil(total_results / 48)
        # print('Total Pages are :', total_pages)

        for page_number in range(1, total_pages + 1):

            # Test
            if page_number == 4:
                return

            offset = page_number * 48
            company = response.meta.get('company', '')

            # Get the Company value from response and make json request for records with pagination
            page_url = f'https://www.cargurus.com/Cars/searchResults.action?entitySelectingHelper.selectedEntity={company}&sourceContext=untrackedExternal_false_0&inventorySearchWidgetType=AUTO&sortDir=ASC&sortType=DEAL_RATING_RPL&shopByTypes=MIX&srpVariation=DEFAULT_SEARCH&nonShippableBaseline=285&offset={offset}&maxResults=48&filtersModified=true'
            yield Request(url=page_url, callback=self.parse_usedcars, dont_filter=True,
                          meta={'url': url}, headers=self.headers)

    def parse_usedcars(self, response):
        # get response and load into json then get records one by one
        try:
            data = json.loads(response.body)

            for car in data:
                item = OrderedDict()
                item['Title'] = car.get('listingTitle', '')
                item['Make'] = car.get('makeName', '')
                item['Model'] = car.get('modelName', '')
                item['Year'] = car.get('carYear', '')
                item['Vin'] = car.get('vin', '')

                item['Location'] = car.get('sellerCity', '')
                item['Condition'] = 'Used'
                item['Price'] = car.get('priceString', '')
                item['Average Market Price'] = f"${car.get('expectedPrice', '')}"
                item['Main Image'] = car.get('originalPictureData', {}).get('url', '')
                item['Options'] = ', '.join([x for x in car.get('options', '')])

                item['Mileage'] = f"{car.get('mileageString', '0')} Miles"
                item['Trim'] = car.get('trimName', '')

                item['Color'] = car.get('exteriorColorName', '') or car.get('normalizedExteriorColor', '')
                item['Body Type'] = car.get('bodyTypeName', '')
                item['Fuel Type'] = car.get('localizedFuelType', '')
                item['Engine'] = car.get('localizedEngineDisplayName', '')

                item['Deal Rating'] = car.get('dealRating', '')
                item['Days on Market'] = car.get('daysOnMarket', '')

                item['Drive Train'] = car.get('driveTrain', '')
                item['City Fuel Economy'] = ' '.join(
                    [str(value) for value in list(car.get('cityFuelEconomy', {}).values())])
                item['Highway Fuel Economy'] = ' '.join(
                    [str(value) for value in list(car.get('highwayFuelEconomy', {}).values())])
                item['Combined Fuel Economy'] = ' '.join(
                    [str(value) for value in list(car.get('combinedFuelEconomy', {}).values())])

                item['Dealer Name'] = car.get('serviceProviderName', '')
                item['Dealer City'] = car.get('sellerCity', '').split(',')[0].strip()
                item['Dealer State'] = car.get('sellerRegion', '')
                item['Dealer Zip'] = car.get('sellerPostalCode', '')
                item['Dealer Phone'] = car.get('phoneNumberString', '')
                item['Dealer Rating'] = round(car.get('sellerRating', 0), 2)
                # item['Dealer Reviews Count'] = car.get('reviewCount', 0)
                item['URL'] = response.meta.get('url') + f"#listing={str(car.get('id', ''))}/NONE/DEFAULT"
                item['Scraped Date'] = self.scraping_datetime

                self.current_scraped_items.append(item)
                self.scraped_product_counter += 1

                yield item

        except Exception as e:
            pass
        
    def find_sold_cars(self):
        if not self.current_scraped_items or not self.previously_scraped_items:
            return

        # Specify the file name and output folder
        sold_output_folder = 'output/Sold'

        # Create the output folder if it doesn't exist
        if not os.path.exists(sold_output_folder):
            os.makedirs(sold_output_folder)

        file_name = f'Sold Cars {datetime.now().strftime("%Y-%m-%d %H%M%S")}.csv'
        sold_cars_filepath = os.path.join(sold_output_folder, file_name)
        master_sold_cars_filepath = os.path.join(sold_output_folder, 'Master SoldCars.csv')

        sold_cars = []

        # Get the list of VINs from current_scraped_items
        current_scraped_vin = [item.get('Vin') for item in self.current_scraped_items]

        # Iterate through previous items, and append to sold_car sheet if VIN not in current_scraped_vin
        for previous_vin, previous_item in self.previously_scraped_items.items():
            if previous_vin not in current_scraped_vin:
                sold_cars.append(previous_item)

        # Save the Sold cars into the new CSV
        self.write_item_into_csv(filename=sold_cars_filepath, rows=sold_cars)

        # Append these sold cars into the master file of sold cars
        self.write_item_into_csv(filename=master_sold_cars_filepath, rows=sold_cars)

        self.mandatory_logs.append(f'{len(sold_cars)} Cars found as Sold Cars as those were exists last time but not exists today')
        self.mandatory_logs.append(f'{len(sold_cars)} Sold Cars written into file: {sold_cars_filepath} as well as added in to the Master Sold cars file: {master_sold_cars_filepath}')

    def write_item_into_csv(self, filename, rows):
        if not rows:
            return

        # Acquire a file lock
        try:
            with self.csv_lock:  # Check with the Lock
                with open(filename, mode='a', newline='', encoding='utf-8') as csv_file:
                    writer = csv.DictWriter(csv_file, fieldnames=self.csv_headers)

                    if csv_file.tell() == 0:
                        writer.writeheader()

                    # Write data to CSV
                    writer.writerows(rows)

        except:
            self.logger.error('Error in Adding Sold cars to Sold File')

    def read_excel_file(self, file_path):
        data = []
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                data.append(row_data)

        except Exception as e:
            print(f"Error reading Excel file: {e}")

        return data

    def get_previous_scarped_cars_from_file(self):
        """
        - Get the cars from the latest previously run file.
        - There could be many files, but we need the latest one to compare the cars
        """

        try:
            rows = []
            output_files = self.get_latest_files_path()

            if not output_files:
                return []

            while True:
                latest_output_filename = output_files.pop(0)

                with open(latest_output_filename, mode='r', encoding='utf-8') as csv_file:
                    rows = list(csv.DictReader(csv_file))

                    if rows:
                        break

                self.mandatory_logs.append(f'Last Scraped file "{latest_output_filename}" has "{len(rows)}" rows to be compared with Current file to find out the Sold Cars')

            return rows

        except:
            self.mandatory_logs.append(f'No previous output file found')
            return []

    def get_latest_files_path(self):
        folder_path = 'output'

        # Get a list of all files in the directory
        files = os.listdir(folder_path)

        # Filter the list to include only CSV files
        csv_files = [f for f in files if f.endswith('.csv')]

        # Sort the CSV files by their last created time (most recent first)
        csv_files.sort(key=lambda x: os.path.getctime(os.path.join(folder_path, x)), reverse=True)

        return [os.path.join(folder_path, csv_file) for csv_file in csv_files] if len(csv_files) > 0 else None

    def write_logs(self):
        if not os.path.exists(self.logs_filename):
            os.makedirs(os.path.dirname(self.logs_filename), exist_ok=True)

        with open(self.logs_filename, mode='a', encoding='utf-8') as logs_file:
            for log in self.mandatory_logs:
                self.logger.info(log)
                # print(log)
                logs_file.write(f'{log}\n')

            logs_file.write(f'\n\n')

        self.mandatory_logs = []

    def close(spider, reason):
        spider.find_sold_cars()

        spider.mandatory_logs.append(f'Total "{spider.scraped_product_counter}" Cars scraped for spider "{spider.name}"')

        spider.mandatory_logs.append(f'Spider "{spider.name}" was started at "{spider.main_start_datetime_str}"')
        spider.mandatory_logs.append(f'Spider "{spider.name}" closed at "{datetime.now().strftime("%Y-%m-%d %H%M%S")}"')

        if spider.error_in_filter:
            spider.mandatory_logs.append(f'\n\nERRORS in spider "{spider.name}"\n')
            spider.mandatory_logs += spider.error_in_filter  # Add all the errors in the logs list and then write into file

        spider.write_logs()
