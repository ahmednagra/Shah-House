import csv
import os
from datetime import datetime
from urllib.parse import parse_qs, urlparse, urljoin

from openpyxl.reader.excel import load_workbook
from scrapy import Spider, Request
from openpyxl import Workbook


def write_to_excel(data, sheet_name):
    # Create the directory if it doesn't exist
    output_dir = 'output/'
    os.makedirs(output_dir, exist_ok=True)

    # Save the workbook
    file_name = f'{output_dir}Booking Names_Price {datetime.now().strftime("%d%m%Y%H%M")}.xlsx'

    # Create a new workbook or load existing workbook if file already exists
    if os.path.isfile(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()

    # Remove the default "Sheet" if it exists
    if "Sheet" in wb.sheetnames:
        sheet = wb["Sheet"]
        wb.remove(sheet)

    # Select the sheet with the spider name or create a new one if it doesn't exist
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(sheet_name)

    existing_data = set(tuple(sheet.iter_rows(values_only=True)))

    headers = ['Name', 'Actual_Price', 'Discounted_Price', 'Date_start', 'Date_end', 'Guests_adult', 'Guest_children',
               'City']

    if sheet.max_row < 1 or tuple(headers) not in existing_data:
        sheet.append(headers)
        existing_data.add(tuple(headers))

    # Write the data rows if they don't already exist
    for row in data:
        flattened_row = [item if not isinstance(item, list) else item[0] for item in row]
        if tuple(flattened_row) not in existing_data:
            sheet.append(flattened_row)
            existing_data.add(tuple(flattened_row))

    try:
        wb.save(file_name)
        print(f"Data saved to {file_name}")
    except Exception as e:
        print(f"An error occurred while saving the data: {str(e)}")


def process_csv_file():
    # Get the absolute path of the current directory
    current_directory = os.path.dirname(os.path.abspath(__file__))

    # Move up two levels to reach the parent directory
    parent_directory = os.path.dirname(current_directory)

    file_path = os.path.join(parent_directory, 'input', 'urls.csv')

    data = []

    try:
        with open(file_path, 'r') as csv_file:
            csv_reader = csv.DictReader(csv_file)
            data = list(csv_reader)

    except FileNotFoundError:
        print(f"File '{file_path}' not found.")

    except Exception as e:
        print(f"An error occurred while reading the file: {str(e)}")

    return data


class NocowanieHomeSpider(Spider):
    name = 'nocowanie_hotels'
    start_urls = ['www.nocowanie.pl']

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.city_urls = [row['nocowanie'] for row in process_csv_file()]

    def start_requests(self):
        for url in self.city_urls:
            yield Request(url=url,
                          callback=self.parse)

    def parse(self, response):
        url_params = parse_qs(urlparse(response.url).query)
        data = []

        all_homes = response.css('article.sr-box')
        for home in all_homes:
            Hotel_Name = home.css('h3.sr-box__title a::text').get('').strip()
            City_Name = url_params.get('q', '')
            Date_start = url_params.get('data[od]', '')
            Date_end = url_params.get('data[do]', '')
            Guests_adult = url_params.get('miejsca_dorosli', '')
            Guest_children = url_params.get('miejsca_dzieci', '')
            Discounted_P = home.css('.price__current::text').get('').replace('zł', '').strip()
            try:
                Discounted_Price = float(Discounted_P)
            except ValueError:
                Discounted_Price = ''
            price_text = home.css('.price__old span::text').get('').replace('\n', '').strip()
            try:
                Actual_Price = float(price_text)
            except ValueError:
                Actual_Price = ''

                # Append the data to the list
            data.append([Hotel_Name, Actual_Price, Discounted_Price, Date_start, Date_end,
                         Guests_adult, Guest_children, City_Name])

        sheet_name = 'nocowanie_hotels'
        write_to_excel(data, sheet_name)

        current_url = response.url
        if 'pid=' not in current_url:
            updated_url = current_url + '&pid=2'
            yield Request(url=updated_url, callback=self.parse)

