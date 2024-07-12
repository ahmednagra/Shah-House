import glob
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scrapy import Request, Spider


class EbayProductsCountSpider(Spider):
    name = "ebay_products_count"
    allowed_domains = ["www.ebay.com"]
    start_urls = ["https://www.ebay.com"]

    custom_settings = {
        'CONCURRENT_REQUESTS': 5,
        'AUTOTHROTTLE_ENABLED': True,
        'AUTOTHROTTLE_START_DELAY': 0.2,
        'AUTOTHROTTLE_MAX_DELAY': 3,

        'RETRY_TIMES': 5,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 404, 408, 429],
    }

    def start_requests(self):
        # Load the Excel file
        wb, ws, filepath = self.load_file()

        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
            # Assuming part number is in the first column
            part_number = row[0].value
            url = f'https://www.ebay.com/sch/i.html?_from=R40&_nkw={part_number}&_sacat=0&rt=nc&LH_Sold=1&LH_Complete=1'
            yield Request(url=url, callback=self.parse,
                          meta={'part_number': part_number, 'row': row, 'row_index': idx})

    def parse(self, response, **kwargs):
        part_no = response.meta['part_number']
        row_index = response.meta['row_index']
        ebay_items_count = response.css('.srp-controls__count-heading .BOLD ::text').get('') or ''

        if ebay_items_count:
            # Load the Excel file
            wb, ws, filepath = self.load_file()

            # Add 'Ebay Items Count' header if not present
            if 'Ebay Items Count' not in [cell.value for cell in ws[1]]:
                self.create_items_count_column(ws)

            # Get the specific row using row_index
            row = list(ws.iter_rows(min_row=row_index, max_row=row_index, min_col=1, max_col=ws.max_column))[0]

            # Check if 'Part Number' matches and 'Ebay Items Count' column is present in the row
            if part_no == row[0].value and 'Ebay Items Count' in [cell.value for cell in ws[1]]:

                # if ebay_items_count_col:
                ebay_items_count_col = self.get_ebay_items_count_col(ws)

                if ebay_items_count_col:
                    # Update the 'Ebay Items Count' for the current row
                    ws.cell(row=row_index, column=ebay_items_count_col, value=ebay_items_count)
                    print(f'Part No: "{part_no}" Successfully Updated ')
                else:
                    self.logger.warning("'Ebay Items Count' column not found in the Excel sheet.")
            else:
                # If part number is not matched, search for it in the entire file and update the specific row
                for excel_row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    if excel_row[0].value == part_no:

                        # Add 'Ebay Items Count' header if not present
                        if 'Ebay Items Count' not in [cell.value for cell in ws[1]]:
                            self.create_items_count_column(ws)

                        ebay_items_count_col = self.get_ebay_items_count_col(ws)
                        if ebay_items_count_col:
                            # Update the 'Ebay Items Count' for the current row
                            ws.cell(row=excel_row[0].row, column=ebay_items_count_col, value=ebay_items_count)
                        else:
                            self.logger.warning("'Ebay Items Count' column not found in the Excel sheet.")

            # Save the updated workbook
            wb.save(filepath)

        else:
            print('No Items Count Found url :', response.url)
            self.logger.warning(f'No Items Count Found url :{response.url}')

    def load_file(self):
        # Load the Excel file
        filepath = glob.glob('*.xlsx')[0]
        wb = load_workbook(filepath)
        ws = wb.active
        return wb, ws, filepath

    def create_items_count_column(self, ws):
        col_letter = get_column_letter(ws.max_column + 1)  # get the Index No for count column
        ws[f'{col_letter}1'] = 'Ebay Items Count'
        return

    def get_ebay_items_count_col(self, ws):
        # Find the column index of 'Ebay Items Count'
        ebay_items_count_col = None
        for cell in ws[1]:
            if cell.value == 'Ebay Items Count':
                ebay_items_count_col = cell.column
                break

        return ebay_items_count_col
