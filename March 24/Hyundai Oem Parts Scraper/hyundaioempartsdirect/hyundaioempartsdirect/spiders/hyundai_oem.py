import os
from datetime import datetime
from urllib.parse import urljoin, unquote
from collections import OrderedDict

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from scrapy import Spider, Request, signals


class HyundaiOemSpider(Spider):
    name = "hyundai_oem"
    start_urls = ["https://www.hyundaioempartsdirect.com"]

    custom_settings = {
        'CONCURRENT_REQUESTS': 1,
        'RETRY_TIMES': 2,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 404, 408, 429]
    }

    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
        'cache-control': 'no-cache',
        'pragma': 'no-cache',
        'priority': 'u=0, i',
        'sec-ch-ua': '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
    }

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.categories = []
        self.category_name = ''
        self.items_scraped_count = 0
        self.total_categories_count = 0
        self.total_products = 0   # Category Found Total Products at pagination
        self.current_scraped_item = []
        self.fitment_scraped_items = []

        # Proxy Scrape OPs:
        self.use_proxy = 'true'
        self.proxy_key = self.read_input_from_file('input/proxy_key.txt')[0]

        self.product_fields = ['Brand', 'SKU', 'Part Number', 'Product Title', 'Category', 'Position',
                               'Image Link1', 'Image Link2', 'Image Link3', 'Image Link4', 'Image Link5',
                               'Description', 'Length (mm)', 'Height', 'Width (mm)', 'Weight (g)', 'Url']

        self.fitment_fields = ['SKU', 'Brand', 'Year', 'Make', 'Model', 'Body & Trim', 'Engine & Transmission', 'Url']
        self.current_dt = datetime.now().strftime("%d%m%Y%H%M%S")
        self.output_directory = f'output/{self.current_dt}'

        self.error = []
        self.mandatory_logs = [f'Spider "{self.name}" Started at "{self.current_dt}"\n']
        self.logs_filename = f'logs/logs {self.current_dt}.txt'

    def start_requests(self):
        yield Request(url=self.start_urls[0], callback=self.parse, headers=self.headers)

    def parse(self, response, **kwargs):
        """
        Parse the initial page to extract Categories and Accessories names and URLs.
        """
        categories = response.css('.cats-list .links-list li') or []
        for category in categories:
            name = category.css('a::text').get('').strip()
            url = category.css('a::attr(href)').get('').strip()

            if name and url:
                brand_info = {
                    'name': name,
                    'url': url
                }

                self.categories.append(brand_info)

        self.total_categories_count = len(self.categories)

    def parse_categories_pagination(self, response):
        products = response.css('.catalog-products .catalog-product')

        self.total_products += int(len(products))
        print(f'Found Products in Category: {self.category_name} and Product :{self.total_products}')
        self.mandatory_logs.append(f'\nCategory: {self.category_name} found total Product :{self.total_products}')

        try:
            for product in products:
                brand = product.css('.product-manufacturer-data strong::text').get('').strip()
                url = ''.join(product.css('.product-title a::attr(href)').get('').split('?')[0:1])

                if url not in self.current_scraped_item:
                    response.meta['brand'] = brand
                    print('product detail : ', urljoin(self.start_urls[0], url))
                    yield Request(url=urljoin(self.start_urls[0], url), callback=self.parse_part_detail, headers=self.headers, meta=response.meta)
                else:
                    self.error.append(f'Product url already Scraped: {url}')
        except Exception as e:
            self.error.append(f'Error Parsing product at Listing page Error:{e}   ::url: {response.url}')

    def parse_part_detail(self, response):
        description = ' '.join(response.css('[itemprop="description"] .description_body ::text').getall())
        sku = response.css('.add-to-cart::attr(data-sku)').get('') or response.css('.sku-display::text').get('')
        part_no = response.css('.sku-display::text').get('').replace('-', '').strip() or response.css('.add-to-cart::attr(data-sku-stripped)').get('')
        brand = response.css('.add-to-cart::attr(data-brand)').get('') or response.meta.get('brand', '')
        product_title = response.css('.product-title-module h1::text').get('').strip()
        images_links = self.get_product_images(response)

        try:
            item = OrderedDict()
            item['Brand'] = brand
            item['SKU'] = sku if sku else ''
            item['Part Number'] = sku
            item['Product Title'] = product_title
            item['Category'] = self.category_name
            item['Position'] = response.css('.positions .list-value ::text').get('').strip()
            item['Description'] = description if description else product_title
            item['Length (mm)'] = ''
            item['Height'] = ''
            item['Width (mm)'] = ''
            item['Weight (g)'] = ''
            item['Url'] = unquote((response.url.split('url=')[1]))

            # Save image links
            for i, img_link in enumerate(images_links[:5]):
                item[f'Image Link{i + 1}'] = img_link

            if not item['SKU'] or not item['Product Title'] or not item['Brand']:
                self.error.append(f'One or more essential fields are empty for URL: {unquote((response.url.split('url=')[1]))}')

            self.items_scraped_count += 1
            print(f'Category {self.category_name} items_scraped_count : {self.items_scraped_count}')
            self.current_scraped_item.append(item)
        except Exception as e:
            self.error.append(f'Parse product Item error: {e} ::URl: {response.url}')

        fitment_table = response.css('.fitment-table tbody tr')
        if fitment_table:
            try:
                for row in fitment_table:
                    fitment_item = OrderedDict()
                    fitment_item['SKU'] = sku
                    fitment_item['Brand'] = brand
                    fitment_item['Year'] = row.css('.fitment-year::text').get()
                    fitment_item['Make'] = row.css('.fitment-make::text').get()
                    fitment_item['Model'] = row.css('.fitment-model::text').get()
                    fitment_item['Body & Trim'] = row.css('.fitment-trim::text').get()
                    fitment_item['Engine & Transmission'] = row.css('.fitment-engine::text').get('')
                    fitment_item['Url'] = unquote((response.url.split('url=')[1]))
                    self.fitment_scraped_items.append(fitment_item)
            except Exception as e:
                self.error.append(f'Parse product Fitment Table error: {e} ::URl: {response.url}')

    def get_product_images(self, response):
        try:
            # Extract All images
            images = [f'https:{img}' for img in
                      response.css('.product-images .product-secondary-image ::attr(data-image-main-url)').getall()]

            # If no All images, try to get main image
            if not images:
                main_image = response.css('.main-image a::attr(href)').get('')
                if main_image:
                    images.append(f'https:{main_image}')

            return images
        except Exception as e:
            print(f"Error occurred while extracting product images: {e}")
            return []

    def write_items_to_excel(self, filename, fields, items):
        """
        Write items to an Excel file.
        """
        os.makedirs(self.output_directory, exist_ok=True)
        output_file = os.path.join(self.output_directory, f'{filename}.xlsx')

        try:
            # Check if file exists
            if os.path.isfile(output_file):
                wb = load_workbook(output_file)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                # Write headers
                for idx, field in enumerate(fields, start=1):
                    ws.cell(row=1, column=idx, value=field)

            # Get the next available row
            next_row = ws.max_row + 1

            # Write data
            for item in items:
                for col_idx, field in enumerate(fields, start=1):
                    cell = ws.cell(row=next_row, column=col_idx)
                    cell.value = item.get(field, '')

                # Increment the row index for the next item
                next_row += 1

            # Save workbook
            wb.save(output_file)

            print(f'{self.category_name} : total items {len(items)}')
        except Exception as e:
            self.error.append(f'Error occurred while writing items to Excel file: "{output_file}", Error: {e}')

    def write_logs(self):
        log_folder = 'logs'
        os.makedirs(log_folder, exist_ok=True)
        with open(self.logs_filename, mode='a', encoding='utf-8') as logs_file:
            for log in self.mandatory_logs:
                self.logger.info(log)
                logs_file.write(f'{log}\n')

            logs_file.write(f'\n\n')

    def read_input_from_file(self, file_path):
        try:
            with open(file_path, mode='r') as txt_file:
                return [line.strip() for line in txt_file.readlines() if line.strip()]

        except FileNotFoundError:
            print(f"File not found: {file_path}")
            self.error.append(f"File not found: {file_path}")
            return []
        except Exception as e:
            print(f"An error occurred: {str(e)}")
            self.error.append(f"An error occurred: {str(e)}")
            return []

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(HyundaiOemSpider, cls).from_crawler(crawler, *args, **kwargs)
        crawler.signals.connect(spider.spider_idle, signal=signals.spider_idle)
        return spider

    def spider_idle(self):
        """
        Handle spider idle state by crawling next brand if available.
        """

        print(f'\n\n{len(self.categories)}/{self.total_categories_count} Categories left to Scrape\n\n')

        if self.current_scraped_item:
            print(f'\n\nTotal {self.items_scraped_count}/ {self.total_products} items scraped from Category: {self.category_name}')
            self.mandatory_logs.append(f'\nTotal {self.items_scraped_count}/ {self.total_products} items scraped from Category: {self.category_name}')
            # Category Product items Csv
            self.write_items_to_excel('Hyundai Oem Parts', self.product_fields, self.current_scraped_item)
            # Category Product Fitments Items to Csv
            self.write_items_to_excel('Hyundai Oem Parts fitments', self.fitment_fields, self.fitment_scraped_items)
            self.category_name = ''
            self.current_scraped_item = []
            self.items_scraped_count = 0
            self.total_products = 0

        if self.categories:
            category_dict = self.categories.pop(0)
            self.category_name = category_dict.get('name', '')
            cat_url = category_dict.get('url', '')

            url = f'https://www.hyundaioempartsdirect.com/{cat_url}'
            req = Request(url=url,
                          callback=self.parse_categories_pagination, headers=self.headers,
                          meta={'handle_httpstatus_all': True})

            try:
                self.crawler.engine.crawl(req)  # For latest Python version
            except TypeError:
                self.crawler.engine.crawl(req, self)  # For old Python version < 10

    def close(spider, reason):
        spider.mandatory_logs.append(f'\nSpider "{spider.name}" was started at "{spider.current_dt}"')
        spider.mandatory_logs.append(f'Spider "{spider.name}" closed at "{datetime.now().strftime("%Y-%m-%d %H%M%S")}"\n\n')

        spider.mandatory_logs.append(f'Spider Error:: \n')
        spider.mandatory_logs.extend(spider.error)

        spider.write_logs()
