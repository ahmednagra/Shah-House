import csv
import os
from datetime import datetime

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

custom_settings = {

    # 'SCRAPEOPS_API_KEY': '69407ad1-67b8-4a4f-8083-137167f3b908',
    # 'SCRAPEOPS_PROXY_ENABLED': True,
    # 'DOWNLOADER_MIDDLEWARES': {
    #     'scrapeops_scrapy_proxy_sdk.scrapeops_scrapy_proxy_sdk.ScrapeOpsScrapyProxySdk': 725,
    # },
    'RETRY_TIMES': 5,
    'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 404, 408],
    'CONCURRENT_REQUESTS': 4,
    'FEED_EXPORTERS': {
        'xlsx': 'scrapy_xlsx.XlsxItemExporter',
    },
    'FEEDS': {
        f'output/Amazon Products Reviews {datetime.now().strftime("%d%m%Y%H%M")}.xlsx': {
            'format': 'xlsx',
            'fields': ['SKU', 'Title', 'Brand', 'Price', 'Availability', 'Size', 'Color', 'URL']
        }
    }
}

def get_scrapeops_url(url):
    payload = {'api_key': API_KEY, 'url': url}
    proxy_url = 'https://proxy.scrapeops.io/v1/?' + urlencode(payload)
    return proxy_url

def read_text_file(self):
    file_path = 'input/city_names.txt'

    try:
        with open(file_path, mode='r') as txt_file:
            return [line.strip() for line in txt_file.readlines() if line.strip()]

    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return []
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return []


def write_to_excel(data, sheet_name):
    # Create the directory if it doesn't exist
    output_dir = 'output/'
    os.makedirs(output_dir, exist_ok=True)

    # Save the workbook
    file_name = f'{output_dir}Booking Names_Price {datetime.now().strftime("%d%m%Y")}.xlsx'

    # Create a new workbook or load existing workbook if file already exists
    if os.path.isfile(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()

    # Select the sheet with the spider name or create a new one if it doesn't exist
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(sheet_name)

    # Write the headers if the sheet is empty
    if sheet.max_row == 0:
        headers = ['Name', 'Actual_Price', 'Discounted_Price', 'Date_start', 'Date_end', 'Guests_adult',
                   'Guest_children', 'City']
        sheet.append(headers)

    # Write the data rows
    for row in data:
        flattened_row = [item if not isinstance(item, list) else item[0] for item in row]
        sheet.append(flattened_row)

    try:
        wb.save(file_name)
        print(f"Data saved to {file_name}")
    except Exception as e:
        print(f"An error occurred while saving the data: {str(e)}")


def process_csv_file(self):
    data = []

    try:
        with open(self.input_file, 'r') as csv_file:
            csv_reader = csv.DictReader(csv_file)
            data = list(csv_reader)

    except FileNotFoundError:
        print(f"File '{self.input_file}' not found.")
        return
    except Exception as e:
        print(f"An error occurred while reading the file: {str(e)}")
        return


def write_csv(date, info_text, case_info):
    # from datetime import datetime

    os.makedirs('output', exist_ok=True)
    filepath = f'output/Case Information {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}'

    # Write data to CSV
    with open(filepath, 'a', newline='') as csvfile:
        writer = csv.writer(csvfile)

        # Check if the file is empty (using tell) and write header if needed
        if csvfile.tell() == 0:
            writer.writerow(['Date', 'Information', 'Case Information'])

        writer.writerow([date, info_text, case_info])

def form_data(self, page_num, category_id):
    return {
        "params": {
            "f_brandno": [],
            "page": page_num,
            "sort": "pageview",
            "f_price": {},
            "category": str(category_id),
            "f_size": [],
            "size": 400,
            "keyword": ""
        },
        "session_id": "71088121b102433db4a2bcdeab689d05",
        "m_no": None
    }


form_params = self.form_data(page_num=1, category_id=category_id)
body = (json.dumps(form_params).encode('utf-8')),
method = 'POST',

# logs
current_dt = datetime.now().strftime("%Y-%m-%d %H%M%S")
os.makedirs('logs', exist_ok=True)
self.logs_filepath = f'logs/logs {self.current_dt}.txt'
self.script_starting_datetime = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
self.write_logs(f'Script Started at "{self.script_starting_datetime}"\n')


def write_logs(self, log_msg):
    with open(self.logs_filepath, mode='a', encoding='utf-8') as logs_file:
        logs_file.write(f'{log_msg}\n')
        print(log_msg)


"""read urls and proxy file same function"""

self.config = self.read_input_file(file_path='input/scrapeops_proxy_key.txt', file_type='config')
self.proxy_key = self.config.get('scrapeops_api_key', '')
self.use_proxy = self.config.get('use_proxy', '')

if spider.use_proxy and 'proxy.scrapeops' not in request.url:
    payload = {'api_key': spider.proxy_key, 'url': request.url}
    proxy_url = 'https://proxy.scrapeops.io/v1/?' + urlencode(payload)
    print('Proxy Url :', proxy_url)
    return request.replace(url=proxy_url)

def read_input_file(self, file_path, file_type):
    """
        Read URLs or configuration from a text file.
        """
    try:
        with open(file_path, mode='r', encoding='utf-8') as file:
            if file_type == 'urls':
                return [line.strip() for line in file.readlines() if line.strip()]
            elif file_type == 'config':
                return {line.split('==')[0].strip(): line.split('==')[1].strip() for line in file if '==' in line}
    except FileNotFoundError:
        self.write_logs(f"File not found: {file_path}")
        return [] if file_type == 'urls' else {}
    except Exception as e:
        self.write_logs(f"An error occurred while reading {file_type} file: {str(e)}")
        return [] if file_type == 'urls' else {}


def get_scrapedo_url(self, url):
    proxy_url = "https://api.scrape.do/?token=17aa66a18dbb40c68bc97b79fbcb94cff3030b22f55&url=" + parse.quote(url)
    return proxy_url



def get_webshare_proxy_list(self, webshare_api_key):

    # using proxy api key from request a variable get the token then from c request variable get the proxies list for specific country
    a = requests.get(
        "https://proxy.webshare.io/api/v2/proxy/config/",
        headers={"Authorization": f"Token {webshare_api_key}"}
        # headers={"Authorization": "Token hjrmxcduqgcisax9huheyg2xwb2xr7qv2nmwa4kj"} # token belong to ahmed
    )
    b = a.json()
    proxy_token = b.get('proxy_list_download_token', '')  # this token mean to download all proxies all country base
    url = f'https://proxy.webshare.io/api/v2/proxy/list/download/{proxy_token}/us/any/username/direct/-/'

    c = requests.get(url=url)
    d = [proxy for proxy in c.text.replace('\r', '').split('\n') if proxy.strip()]

    # from this req variable get the all proxies in list  with pagination option.
    req = requests.get(
        "https://proxy.webshare.io/api/v2/proxy/list/?mode=direct&page=1&page_size=250",
        headers={"Authorization": f"Token {webshare_api_key}"}
    )

    proxy_list = []
    for proxy in d:
        proxy = proxy.split(':')
        ip = proxy[0]
        port_no = proxy[1]
        username = proxy[2]
        password = proxy[3]
        new_proxy = f"http://{username}:{password}@{ip}:{port_no}"
        print('New Proxy Url :', new_proxy)
        proxy_list.append(new_proxy)

    return proxy_list


proxy = random.choice(self.proxy_list)
host, port, user, password = proxy.split(':')
proxy = f"http://{user}:{password}@{host}:{port}"


