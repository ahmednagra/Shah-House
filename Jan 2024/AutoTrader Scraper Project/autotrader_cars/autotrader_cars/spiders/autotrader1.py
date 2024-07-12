import csv
import glob
import json
import os
from collections import OrderedDict
from datetime import datetime
from time import sleep

import requests
from scrapy import Spider, Request, signals

from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook


class AutotraderSpider(Spider):
    name = 'autotrader'
    base_url = 'https://www.autotrader.com/'

    csv_headers = ['Title', 'Make', 'Model', 'Year', 'Vin',
                   'Location', 'Condition', 'Price', 'Average Market Price',
                   'Status', 'Main Image', 'Options', 'Mileage', 'Trim', 'Color',
                   'Body Type', 'Fuel Type', 'Engine', 'Drive Train', 'City Fuel Economy',
                   'Highway Fuel Economy', 'Combined Fuel Economy', 'Dealer Name',
                   'Dealer City', 'Dealer State', 'Dealer Zip', 'Dealer Rating',
                   'Dealer Phone', 'URL']

    custom_settings = {

        'RETRY_TIMES': 5,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 404, 408],
        'AUTOTHROTTLE_ENABLED': True,
        'AUTOTHROTTLE_START_DELAY': 1,
        'AUTOTHROTTLE_MAX_DELAY': 3,

        'CONCURRENT_REQUESTS': 4,
        'DOWNLOAD_DELAY': 1,
    }

    headers = {
        'authority': 'www.autotrader.com',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
        'cache-control': 'no-cache',
        'pragma': 'no-cache',
        'sec-ch-ua': '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'none',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    }

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # self.output_filename = f'output/Autotrader Used Cars {datetime.now().strftime("%d%m%Y%H")}.csv'
        self.output_filename = f'output/Autotrader Used Cars {datetime.now().strftime("%d%m%Y%H%M%S")}.csv'
        filers = self.get_key_values_from_file()
        years = filers.get('year', '').split(',') if filers.get('year') else ['0', '0']  # avoid from error if year emty in file
        self.make_filter = [maker.lower() for maker in filers.get('make', '').split(',')]
        self.model_filter = [model.replace('-', '').lower().strip() for model in filers.get('model', '').split(',')]

        self.year_range_filter = [str(year) for year in range(int(''.join(years[:1])), int(''.join(years[1:2]))+1) if not year == 0]

        self.start_time = datetime.now().strftime('%d-%m-%Y %H-%M')
        self.error_in_filter = []
        self.need_filters = []
        self.skipped_cars = 0
        self.makers = None
        self.blocked_urls = []
        self.current_scraped_items_p_ids = []
        self.current_scraped_items = []
        self.scraped_product_counter = 0
        self.previously_scraped_items = {item.get('Vin'): item for item in self.read_scraped_cars()}
        self.master_file_rows = {item.get('Vin'): item for item in self.read_excel_file('output/Sold/Master SoldCars.xlsx')}
        d=1

    def start_requests(self):
        yield Request(url=self.base_url, headers=self.headers)

    def parse(self, response, **kwargs):
        # makers = response.css('[label="All Makes"] option ::attr(value)').getall()
        makers = response.css('[label="All Makes"] option')

        makers = [maker.css('::attr(value)').get('') for maker in makers if maker.css('::attr(label)').get('').lower() not in self.make_filter]

        self.makers = makers[:35]

    def parse_maker_cars(self, response):
        # after writing the current scraped items into file now empty the list container
        self.current_scraped_items = []

        maker = self.makers.pop(0)

        url = f'https://www.autotrader.com/rest/lsc/listing?makeCode={maker}&firstRecord=0&newSearch=false&numRecords=2000&listingType=USED'
        yield Request(url=url, headers=self.headers, callback=self.parse_car_maker, meta={'make': maker})

    def parse_car_maker(self, response):
        make = response.meta.get('make')
        json_data, total_results = self.get_json_data(response)

        self.logger.info(f'\n\nMaker: {make} have Products: {total_results}\n')

        if total_results > 2000:
            url = f'https://www.autotrader.com/collections/ccServices/rest/ccs/models?makeCode={make}'
            yield Request(url=url, headers=self.headers, callback=self.parse_car_maker_models,
                          meta={'search_url': response.url})
        else:
            yield from self.parse_products(json_data)

    def parse_car_maker_models(self, response):
        json_data, total_results = self.get_json_data(response)

        try:
            # maker_models = [model.get('value') for model in json_data if model.get('value')]
            maker_models = {model.get('label', '').replace('-', '').strip().lower(): model.get('value') for model in json_data if model.get('value')}
        except:
            maker_models = {}

        for model_label, model_value in maker_models.items():
            # apply the input filters
            if model_label in self.model_filter:
                self.logger.info(f'\n\nSkipped {model_label} Model\n')
                continue

            url = f'{response.meta.get("search_url")}&modelCode={model_value}'
            yield Request(url=url, headers=self.headers, callback=self.parse_car_year_range)

    def parse_car_year_range(self, response):
        json_data, total_results = self.get_json_data(response)

        if total_results > 2000:
            try:
                year_range = [year.get('value') for year in json_data.get('filters', {}).get('yearRange', {}).get('options') if year.get('value')]
            except:
                self.error_in_filter.append(response.url)
                year_range = []
                yield from self.parse_products(json_data)

            for year in year_range:

                if year == '0':
                    continue

                if year in self.year_range_filter:
                    continue

                yield Request(url=f'{response.url}&endYear={year}&startYear={year}', headers=self.headers,
                              callback=self.parse_car_colors)
        else:
            yield from self.parse_products(json_data)

    def parse_car_colors(self, response):
        json_data, total_results = self.get_json_data(response)

        if total_results > 2000:
            try:
                # colors = [color.get('value') for color in json_data.get('filters', {}).get('intColorSimple', {}).get('options', [{}])]
                colors = [color.get('value') for color in json_data.get('filters', {}).get('extColorSimple', {}).get('options', [{}])]
            except:
                self.error_in_filter.append(response.url)
                colors = []
                yield from self.parse_products(json_data)

            for color in colors:
                # url = f"{response.url}&intColorSimple={color}"
                url = f"{response.url}&extColorSimple={color}"
                yield Request(url=url, headers=self.headers, callback=self.parse_car_body_styles)
        else:
            yield from self.parse_products(json_data)

    def parse_car_body_styles(self, response):
        json_data, total_results = self.get_json_data(response)

        if total_results > 2000:
            try:
                body_styles = [option.get('value') for option in json_data.get('filters', {}).get('vehicleStyleCode', {}).get('options')]
            except:
                self.error_in_filter.append(response.url)
                body_styles = []
                yield from self.parse_products(json_data)

            for style in body_styles:
                url = f"{response.url}&vehicleStyleCode={style}"
                yield Request(url=url, headers=self.headers, callback=self.parse_car_fuel_type)

        else:
            yield from self.parse_products(json_data)

    def parse_car_fuel_type(self, response):
        json_data, total_results = self.get_json_data(response)

        if total_results > 2000:
            try:
                fuel_types = [option.get('value') for option in json_data.get('filters', {}).get('fuelTypeGroup', {}).get('options')]
            except:
                self.error_in_filter.append(response.url)
                fuel_types = []
                yield from self.parse_products(json_data)

            for f_type in fuel_types:
                url = f"{response.url}&fuelTypeGroup={f_type}"
                yield Request(url=url, headers=self.headers, callback=self.parse_car_engine_displacement)

        else:
            yield from self.parse_products(json_data)

    def parse_car_engine_displacement(self, response):
        json_data, total_results = self.get_json_data(response)
        if total_results > 2000:

            engine_displacement = [option.get('value') for option in json_data.get('filters', {}).get('engineDisplacement', {}).get('options')]
            for engine_dis in engine_displacement:
                url = f"{response.url}&engineDisplacement={engine_dis}"
                yield Request(url=url, headers=self.headers, callback=self.parse_cars)

        else:
            yield from self.parse_products(json_data)

    def parse_cars(self, response):
        json_data, total_results = self.get_json_data(response)
        if total_results > 2000:
            for sort_option in ['derivedpriceASC', 'derivedpriceDESC']:
                url = f"{response.url}&sortBy={sort_option}"
                yield Request(url=url, headers=self.headers, callback=self.parse_sorted_cars)
        else:
            yield from self.parse_products(json_data)

    def parse_sorted_cars(self, response):
        json_data, total_results = self.get_json_data(response)
        yield from self.parse_products(json_data)

    def parse_products(self, json_data):
        products = json_data.get('listings', [])

        if not products:
            return

        for product in products:
            try:
                # Extract basic information from the JSON data
                p_id = product.get('id', '')
                if p_id in self.current_scraped_items_p_ids:
                    self.skipped_cars += 1
                    # print(f'{p_id} already Exist')
                    continue

                make = product.get('make', {}).get('name', '')
                model = product.get('model', {}).get('name', '')
                year = str(product.get('year', 0))

                model = model.replace(str(make), '').replace(year, '').strip()

                if make.lower() in self.make_filter:
                    continue

                if model.replace('-', '').strip().lower() in self.model_filter:
                    continue

                if year in self.year_range_filter:
                    continue

                condition = product.get('listingTypes', {})
                if condition:
                    condition = product.get('listingTypes', {})[0].get('name', '') or product.get('listingType', '')
                price = product.get('pricingDetail', {}).get('salePrice', 0)
                avg_market_price = product.get('pricingDetail', {}).get('kbbFppAmount', 0.0)
                milage = product.get('mileage', {}).get('value', '').strip()

                # Extract specific specifications
                body_styles = product.get('bodyStyles', [])
                if body_styles:
                    body_type = body_styles[0].get('code', '')
                else:
                    body_type = ''

                city_fuel = product.get('mpgCity', 0)
                highway_fuel = product.get('mpgHighway', 0)

                # Create an ordered dictionary to store item data
                item = OrderedDict()
                item['Title'] = product.get('title', '').replace(year, '').replace(condition, '').strip()
                item['Make'] = make
                item['Model'] = model
                item['Year'] = year if year else ''
                item['Vin'] = product.get('vin', '')

                # Populate additional fields
                item['Location'] = ''
                item['Condition'] = condition
                item['Price'] = price if price else ''
                item['Average Market Price'] = avg_market_price if avg_market_price else ''
                item['Main Image'] = ''.join([img.get('src', '') for img in product.get('images', {}).get('sources', [{}])][:1])
                item['Options'] = self.get_options(product)
                item['Mileage'] = milage if milage else ''
                item['Trim'] = product.get('trim', {}).get('code', '')
                item['Color'] = product.get('color', {}).get('exteriorColor', '')
                item['Body Type'] = body_type
                item['Fuel Type'] = product.get('fuelType', {}).get('name', '')
                item['Engine'] = product.get('engine', {}).get('name', '')
                item['Drive Train'] = product.get('driveType', {}).get('description', '')
                item['City Fuel Economy'] = city_fuel if city_fuel else ''
                item['Highway Fuel Economy'] = highway_fuel if highway_fuel else ''
                item['Combined Fuel Economy'] = ''

                # Extract dealer information
                dealer_address = product.get('owner', {}).get('location', {}).get('address', {})
                dealer_rating = product.get('owner', {}).get('rating', {}).get('value', 0.0)
                item['Dealer Name'] = product.get('owner', {}).get('name', '')
                item['Dealer City'] = dealer_address.get('city', '').strip()
                item['Dealer State'] = dealer_address.get('state', '')
                item['Dealer Zip'] = dealer_address.get('zip', '')
                item['Dealer Rating'] = dealer_rating if dealer_rating else ''
                item['Dealer Phone'] = product.get('owner', {}).get('phone', {}).get('value', '')
                item['URL'] = f'https://www.autotrader.com/cars-for-sale/vehicle/{p_id}'
                item['Status'] = product.get('pricingDetail', {}).get('dealIndicator', '')
                item['Scraped Date'] = datetime.now().strftime('%Y-%m-%d')

                self.scraped_product_counter += 1
                print('Current Scraped Items Counter :', self.scraped_product_counter)
                self.current_scraped_items_p_ids.append(p_id)

                # self.write_item_into_csv_file(item)
                self.current_scraped_items.append(item)

                yield item

            except:
                continue

    def write_item_into_csv_file(self):

        if not self.current_scraped_items:
            return

        fieldnames = self.current_scraped_items[0].keys()

        with open(self.output_filename, mode='a', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            if csvfile.tell() == 0:
                writer.writeheader()

            for item in self.current_scraped_items:
                writer.writerow(item)

    def get_options(self, product):
        specifications = product.get('specifications', {})
        options_string = ""

        for key, value in specifications.items():
            label = value.get('label', '')
            option_value = value.get('value', '')
            option_string = f"{label}: {option_value}"
            options_string += f"{key.title()} = [{option_string}]\n"

        return options_string.strip()

    def get_json_data(self, response):
        try:
            json_data = json.loads(response.text)
        except:
            # resp = requests.get(response.url, headers=self.headers)
            resp = self.make_request_with_retry(response.url)

            try:
                json_data = json.loads(resp.text)
            except:
                json_data = {}
                self.blocked_urls.append(response.url)

        try:
            total_results = json_data.get('totalResultCount', 0)
        except:
            total_results = 0

        return json_data, total_results

    def get_key_values_from_file(self):
        with open(r'input\filters.txt', mode='r', encoding='utf-8') as input_file:
            data = {}

            for row in input_file.readlines():
                if not row.strip():
                    continue

                try:
                    key, value = row.strip().split('==')
                    data.setdefault(key.strip(), value.strip())
                except ValueError:
                    pass

            return data

    def make_request_with_retry(self, url):
        max_retries = 3

        for attempt in range(1, max_retries + 1):
            try:
                response = requests.get(url, headers=self.headers)
                response.raise_for_status()
                return response
            except requests.exceptions.RequestException as e:
                # print(f"Attempt {attempt} failed. Retrying in 2 seconds. Error: {e}")
                sleep(2)

        raise Exception(f"Max retry attempts reached for URL: {url}")

    def read_scraped_cars(self):

        try:
            latest_input_file = self.get_latest_file_path()
            with open(latest_input_file, mode='r', encoding='utf-8') as csv_file:
                return list(csv.DictReader(csv_file))
        except:
            return []

    def get_latest_file_path(self):
        folder_path = 'output'
        files = [file for file in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, file))]
        latest_file = max(files, key=lambda x: os.path.getmtime(os.path.join(folder_path, x)))
        return os.path.join(folder_path, latest_file)

    def read_excel_file(self, file_path):
        data = []
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                data.append(row_data)

        except Exception as e:
            print(f"Error reading Excel file: {e}")

        return data

    def get_work_sheet(self, sheet_name, wb, headers):
        # function for worksheet select or name the worksheet
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)
            sheet.append(headers)

        return sheet

    def find_sold_cars(self):
        if not self.current_scraped_items or not self.previously_scraped_items:
            return

        current_time = datetime.now()
        formatted_time = current_time.strftime("%Y%m%d %H%M%S")

        # Specify the file name and output folder
        output_folder = 'output/Sold'
        file_name = f'SoldCars {formatted_time}.xlsx'
        sold_cars_filepath = os.path.join(output_folder, file_name)

        # Create the output folder if it doesn't exist
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Create or load the workbook
        if os.path.isfile(sold_cars_filepath):
            wb = load_workbook(sold_cars_filepath)
        else:
            wb = Workbook()

        if "Sheet" in wb.sheetnames:
            sheet = wb["Sheet"]
            wb.remove(sheet)

        # # Create or get sheets
        sold_cars = self.get_work_sheet(sheet_name='Sold Cars', wb=wb, headers=self.csv_headers)

        # Get the list of VINs from current_scraped_items
        current_scraped_vin = [item.get('Vin') for item in self.current_scraped_items]
        sold_cars_count = 0

        # Iterate through previous items, and append to sold_car sheet if VIN not in current_scraped_vin
        for previous_vin, previous_item in self.previously_scraped_items.items():
            if previous_vin not in current_scraped_vin:
                row_values = list(previous_item.values())
                sold_cars.append(row_values)
                self.master_file_rows.update({previous_vin: previous_item.values()})
                sold_cars_count += 1
                a=0

        # Save the workbook
        wb.save(sold_cars_filepath)
        self.logger.info(f'{sold_cars_count} Sold Cars written into file: {sold_cars_filepath}')

    def write_master_sold_car_file(self):
        output_folder = 'output/Sold'
        master_file_sold_cars = f'Master SoldCars.xlsx'
        master_sold_cars_filepath = os.path.join(output_folder, master_file_sold_cars)

        # Create or load the workbook master file
        if os.path.isfile(master_sold_cars_filepath):
            master_wb = load_workbook(master_sold_cars_filepath)
        else:
            master_wb = Workbook()

        if "Sheet" in master_wb.sheetnames:
            sheet = master_wb["Sheet"]
            master_wb.remove(sheet)

        master_sold_cars = self.get_work_sheet(sheet_name='Sold Cars', wb=master_wb, headers=self.csv_headers)

        for item in self.master_file_rows.values():
            master_sold_cars.append(item)

        master_wb.save(master_sold_cars_filepath)


    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(AutotraderSpider, cls).from_crawler(crawler, *args, **kwargs)
        crawler.signals.connect(spider.spider_idle, signal=signals.spider_idle)
        return spider

    def spider_idle(self):
        # after every maker complete scraped then write scraped items into csv file
        self.write_item_into_csv_file()

        if self.makers:
            self.crawler.engine.crawl(Request(url=self.base_url,
                                              callback=self.parse_maker_cars,
                                              dont_filter=True,
                                              headers=self.headers,
                                              meta={'handle_httpstatus_all': True}))

    def close(spider, reason):
        spider.find_sold_cars()
        spider.write_master_sold_car_file()

        spider.logger.info(f'\n Scraped Cars: {spider.scraped_product_counter}')
        spider.logger.info(f'\n Skipped Cars: {spider.skipped_cars}')

        spider.logger.info(f"\n\nScript Start time: {spider.start_time}")
        spider.logger.info(f"\nScript close time: {datetime.now().strftime('%d-%m-%Y %H-%M')}")


