import os
import json
from datetime import datetime
from collections import OrderedDict

from openpyxl import load_workbook
from scrapy import Spider, Request


class Allabolag1Spider(Spider):
    name = 'allabolag'
    base_url = 'https://www.allabolag.se'
    current_datetime = datetime.now().strftime('%d-%m-%Y %H%M%S')
    # start_urls = ['https://www.allabolag.se/branscher/bank-finans-forsakring/14']
    # start_urls = ['https://www.allabolag.se/branscher/bemanning-arbetsformedling/23']
    start_urls = ['https://www.allabolag.se/branscher/juridik-ekonomi-konsulttjanster/16']  # Category: "Juridik, Ekonomi & Konsulttjänster"

    custom_settings = {
        'CONCURRENT_REQUESTS': 5,

        'RETRY_TIMES': 5,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 402, 403, 404, 408, 429, 484, 10051],

        'AUTOTHROTTLE_ENABLED': 'True',
        'AUTOTHROTTLE_START_DELAY': 0.1,
        'AUTOTHROTTLE_MAX_DELAY': 5,

        'FEED_EXPORTERS': {
            'xlsx': 'scrapy_xlsx.XlsxItemExporter',
        },

        'FEED_FORMAT': 'xlsx',
        # 'FEED_URI': f'output/Companies {current_datetime}.xlsx',
        'FEED_URI': f'output/Juridik-Ekonomi-Konsulttjanster {current_datetime}.xlsx',
        'FEED_EXPORT_FIELDS': ['Organization Number', 'Company Name', 'Company Address', 'Significant Person', 'Revenue', 'Url']
    }

    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
        'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Mobile Safari/537.36',
    }

    json_headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
    }

    def __init__(self):
        super().__init__()
        self.companies_urls = []
        self.companies_count = 0
        self.total_companies_count = 0
        self.duplicate_companies_count = 0

        os.makedirs('custom_logs', exist_ok=True)
        os.makedirs('running_logs', exist_ok=True)
        self.logs_filepath = f'custom_logs/logs {self.current_datetime}.txt'
        self.script_starting_datetime = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        self.write_logs(f'Script Started at "{self.script_starting_datetime}"\n')

        # previous_companies_scraped = self.get_scraped_companies_from_file('output/Companies 14-06-2024 230608.xlsx')
        # self.previous_scraped_companies_urls = [item.get('Url') for item in previous_companies_scraped]

    def start_requests(self):
        yield Request(url=self.start_urls[0], headers=self.headers, callback=self.parse)

    def parse(self, response, **kwargs):
        """Parse the categories on the start page and send requests to each category."""
        categories_selectors = response.css('.tw-list-disc li a')

        for category in categories_selectors:
            category_name = ''.join(category.css('a::text').get('').split('(')[0:1]).strip()
            category_url = category.css('a::attr(href)').get('')
            url = f'{self.base_url}{category_url}'
            yield Request(url=url, headers=self.json_headers, callback=self.parse_pagination, meta={'category_name': category_name})

    def parse_pagination(self, response):
        """Handle pagination and filter results by country if necessary."""
        data = self.get_data(response)
        total_companies = data.get('totalHits', 0)
        category_name = response.meta.get('category_name', '')
        country_name = response.meta.get('country_name', '')

        if total_companies < 8000:
            if country_name:
                self.write_logs(f'{category_name} , Country : {country_name}  Has Total Companies : {total_companies}')
            else:
                self.write_logs(f'{category_name} Has Total Companies : {total_companies}')

            self.total_companies_count += int(total_companies)

        # Country Filter
        if 'search_index' not in response.meta.get('search_index', '') and total_companies > 8000:
            country_filter_dict = data.get('allFiltersVue', {}).get('xl', {})
            for key in country_filter_dict:
                if key.isdigit():
                    search_index = country_filter_dict[key].get('searchIndex', '')
                    country_name = country_filter_dict[key].get('name', '')
                    if search_index:
                        url = f'{response.url}/{search_index}'
                        response.meta['search_index'] = search_index
                        response.meta['country_name'] = country_name
                        yield Request(url=url, headers=self.json_headers, callback=self.parse_pagination, meta=response.meta)
            return

        # pagination
        last_page = data.get('paginator', {}).get('last_page', 0)

        for page_number in range(1, int(last_page) + 1):
            yield Request(
                url=f'{response.url}?page={page_number}',
                headers=self.json_headers,
                callback=self.listing_items,
                dont_filter=True,
            )

    def listing_items(self, response):

        """Parse the listing items on a page and send requests to each company's detail page."""
        data = self.get_data(response)
        companies_results = data.get('hitlistVue', [])
        companies_urls = [url.get('linkTo', '') for url in companies_results]

        for company_url in companies_urls:
            url = f'{self.base_url}/{company_url}'

            # url = 'https://www.allabolag.se/2021002684/sveriges-riksbank'

            # if url in self.previous_scraped_companies_urls:
            #     continue

            if url in self.companies_urls:
                self.write_logs(f"Url already scraped : {url}")
                self.duplicate_companies_count += 1
                continue

            self.companies_urls.append(url)

            yield Request(url=url, headers=self.headers, callback=self.parse_company_detail)

    def parse_company_detail(self, response):
        """Parse the details of a single company."""
        try:
            revenue = response.css('th:contains("Omsättning") + td ::text ').get('').strip() or response.css('dt:contains("OMSÄTTNING") + dd::text').get('')

            item = OrderedDict()
            item['Organization Number'] = response.css('.orgnr::text').re_first(r'([0-9-X]+)')
            item['Company Name'] = response.css('.tablet\:tw-text-2xl::text').get('')
            item['Company Address'] = self.get_company_address(response)
            item['Significant Person'] = response.css('.tw-mb-4 dd:nth-child(2) .btn-link::text, dt:contains("CO") + dd ::text').get('')
            item['Revenue'] = '' if revenue.strip() == '0' else revenue.strip()
            item['Url'] = response.url

            self.companies_count += 1
            print(f'\n\nItems Scraped Count: {self.companies_count}\n\n')
            yield item
        except Exception as e:
            self.write_logs(f'Parsing Detail error :{e}, Url: {response.url}')

    def get_data(self, response):
        """Parse JSON data from the response."""
        try:
            return response.json()
        except json.JSONDecodeError as e:
            self.write_logs(f'Error parsing JSON response: {e}')
            return {}

    def write_logs(self, log_msg):
        with open(self.logs_filepath, mode='a', encoding='utf-8') as logs_file:
            logs_file.write(f'{log_msg}\n')
            print(log_msg)

    def get_significant_person(self, response):
        """Extract the significant person's name from the response."""
        person_name = response.css('.tw-mb-4 dd:nth-child(2) .btn-link::text, dt:contains("CO") + dd ::text').get('')
        return person_name

    def get_company_address(self, response):
        """Extract the company address from the response."""
        # Extract Ort and Län
        place_country = [
            response.css(f'dt:contains("{label}") + dd::text').get(default='').strip()
            for label in ["Ort", "Län"]
        ]
        place_country = ', '.join(filter(None, place_country))

        # Extract Besöksadress
        visiting_address = ', '.join([
            text.strip() for text in response.css('dt:contains("Besöksadress") + dd::text, dt:contains("Utdelningsadress") + dd::text').getall() if text.strip()
        ])

        # Combine visiting address and place country
        if visiting_address or place_country:
            address = f"{visiting_address}, {place_country}".strip(', ')
        else:
            address = response.css('#company-card_overview .desktop-only::text').get(default='').strip()

        return address

    def get_scraped_companies_from_file(self, filename):
        data = []
        # print(f'\n\nGetting data from Excel file: "{filename}"')

        workbook = load_workbook(filename)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)

        # print(f'{len(data)} Rows fetched')
        return data


    def close(spider, reason):
        """Log the closing information of the spider."""
        spider.write_logs(f'\n\nFrom Total Companies {spider.total_companies_count} , Companies are scraped:{spider.companies_count}')
        spider.write_logs(f'{spider.name} Total Duplicate Company found : {spider.duplicate_companies_count}')
        spider.write_logs(f'\n\nSpider "{spider.name}" was started at "{spider.current_datetime}"')
        spider.write_logs(f'Spider "{spider.name}" closed at "{datetime.now().strftime("%Y-%m-%d %H%M%S")}"')
