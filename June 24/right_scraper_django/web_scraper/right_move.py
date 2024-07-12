import glob

import requests
import json
from datetime import datetime
from collections import OrderedDict
from parsel import Selector
from openpyxl.workbook import Workbook
import os


# Function to read property URLs from a file

def get_property_urls_from_file():
    try:
        current_dir = os.getcwd()
        parent_dir = os.path.dirname(current_dir)
        input_folder_path = os.path.join(parent_dir, 'input')

        txt_files = glob.glob(os.path.join(input_folder_path, '*.txt'))
        file_path = ''.join([file for file in txt_files if 'property_urls.txt' in file][0:1])
        with open(file_path, 'r', encoding='utf-8') as file:
            urls = file.readlines()
        return [url.strip() for url in urls if url.strip()]
    except:
        print('File Not exist')
        return 'https://www.rightmove.co.uk/properties/150032027#/?channel=OVERSEAS'


# Function to write data to an Excel file
def write_to_excel(data):
    current_dir = os.getcwd()
    parent_dir = os.path.dirname(current_dir)
    # Form the path to the 'input' folder
    output_folder_path = os.path.join(parent_dir, 'output')

    output_file = os.path.join(output_folder_path,
                               f'RightMove Properties {datetime.now().strftime("%d%m%Y %H%M%S")}.xlsx')

    wb = Workbook()
    sheet = wb.active

    # Write the headers if the sheet is empty
    if sheet.max_row == 0 or sheet.max_row == 1:
        # headers = list(data[0].keys())
        headers = ['Address', 'Price PCM', 'Price PW', 'Property Type', 'Bedrooms', 'Bathrooms', 'Available Date',
                   'Furnish Type', 'image_urls', 'Floor Plan image url']
        sheet.append(headers)

    # Write the data rows
    for row in data:
        row_values = []
        for header in headers:
            if isinstance(row[header], list):
                row_values.append(', '.join(row[header]))
            else:
                row_values.append(row.get(header, ''))
        sheet.append(row_values)

    try:
        wb.save(output_file)
        print(f"Data saved to {output_file}")
    except Exception as e:
        print(f"An error occurred while saving the data: {str(e)} data {data}")


# Function to get value by heading
def get_value_by_heading(selector, heading, letting_details=False):
    if letting_details:
        return selector.css(f'dt:contains("{heading}") + dd::text').get(default='').strip()
    else:
        return selector.css(f'dl:contains("{heading}") dd::text').get(default='').strip()


# Function to get images from the response
def get_images(selector):
    try:
        json_data = json.loads(
            selector.css('script:contains("propertyData")::text').re_first(r'window.PAGE_MODEL = (.*)'))
        property_json = json_data.get('propertyData', {}) or {}
        property_images = [image.get('url') for image in property_json.get('images', [{}])]
        floor_plan_image = property_json.get('floorplans', [{}])[0].get('url') or None
    except (json.JSONDecodeError, AttributeError, IndexError):
        floor_plan_image = None
        property_images = []

    floor_plan = floor_plan_image or selector.css('a[href*="plan"] img::attr(src)').get('').replace('_max_296x197', '')
    image_urls = property_images or selector.css('a[itemprop="photo"] [itemprop="contentUrl"]::attr(content)').getall()

    images = {f'Image {index}': '' for index in range(1, 11)}

    images.update({f'Image {index + 1}': f'=IMAGE("{image_url}")' for index, image_url in enumerate(image_urls)})
    images.update({'Floor Plan': f'=IMAGE("{floor_plan}")'})

    return images, image_urls, floor_plan


# Main function to scrape data
def main(url):
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        # 'Cookie': 'permuserid=240629OVRKND5ANQCKSMUVEHHPA0WK6J; beta_optin=N:60:-1; RM_Register=C; OptanonAlertBoxClosed=2024-06-29T07:41:25.490Z; eupubconsent-v2=CQA9lHgQA9lHgAcABBENA7EsAP_gAEPgAChQJhNB9C7dbWFDcH53YLtkMA0HwVAI4sQgBgCAE-IFDBKUIIQGgmAQJASgBCACAAIAIAZBIAMkGAAACUAAYIABCABMAAEEIAAAIAAAAAABAgAIAAACAAEAEAAIgEAQEkAAmAgFAJIASEwAhAAAAAAAACAEAIABAgAAAAAAQAAIAAAIQCgAAAAAAAAAAAAAQBAAAAAAAAAAAAAAAQP3ADIFCogAKIkJCCQEIIAAIggCAAgQAAAAEABAQAGCAoQAgEIMAEAAAAAAAAAAAAQAACAAAAABAAAIAAgQAAACAQAAAAQAAAAIAAAAACAAAAAAAEAAAAAAAAAAAEAIABAhCAACAAkIAAAABAAAAAAUAAAQAAEAAAAAAAAAAAAAAQAA.f_wACHwAAAAA; OTAdditionalConsentString=1~70.89.196.311.494.864.1097.1188.1364.1423.1570.1659.1870.1985.1987.2103.2213.2343.2477.2526.2577.2605.2661.2677.2714.2821.2869.2901.2947.3100.3126.3253.3309.14332.15731.16931.21233.23031.24431.25731.25931.26031.26831.27731.27831.28031.28731.28831.29631.31631; TS01ec61d1=012f990cd3e219a41146a1f50b4fa3824834eaf347f331f0851562741e5834e187b340f0c44217649596f019f540f4809bca06f423; TS01826437=012f990cd3d8c560058356ccdf47d6188320e97031c6eb9199932d7ea414e656693135424e127fe7cfdacbe48a354e7e45f27d420e; ppid=240629OVRKND5ANQCKSMUVEHHPA0WK6J; TPCminPrice=2750; TPCmaxPrice=4000; TS01821201=012f990cd38b9ef4fbffbfef6f33cbb7f7366958179d4d642b717b37c399bac40fab8bf7635808bd6b9e457e0e7ddbed2b8129e771; JSESSIONID=C3093796C0914749D54901397D0B3A13; svr=3114; lastSearchChannel=RES_OVS; lastViewedFeaturedProperties=149525348|143805812|149964932; rmsessionid=c3f37a65-e411-4a14-ba3f-a0a47c06d211; TS019c0ed0=012f990cd3c05ef194c4ad92c4a9ccc9dac277200719e5794141663f6097fa48348bf3f49af66e34947b7c52730199f537837a5aca; OptanonConsent=isGpcEnabled=0&datestamp=Tue+Jul+09+2024+13%3A18%3A51+GMT%2B0500+(Pakistan+Standard+Time)&version=202406.1.0&browserGpcFlag=0&isIABGlobal=false&hosts=&landingPath=NotLandingPage&groups=1%3A1%2C3%3A1%2C4%3A1%2CV2STACK42%3A1&geolocation=%3B&AwaitingReconsent=false',
        'Pragma': 'no-cache',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'none',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Not/A)Brand";v="8", "Chromium";v="126", "Google Chrome";v="126"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }
    data = []

    try:
        print('Requested URL:', url)
        response = requests.get(url, headers=headers)

        if response.status_code == 410:
            print(f'URL {url} returned 410 status code')
            return data

        if response.status_code != 200:
            print(f'URL {url}  returned response code {response.status_code}')
            return data

        selector = Selector(response.text)
        try:
            json_data = json.loads(
                selector.css('script:contains("propertyData")::text').re_first(r'window.PAGE_MODEL = (.*)')).get(
                'propertyData', {})
        except:
            json_data = {}

        images, images_urls, floor_plan = get_images(selector)

        item = OrderedDict()
        item['Address'] = selector.css('[itemprop="streetAddress"]::text').get(default='').strip()
        item['Price PCM'] = selector.css('article div span:contains(" pcm")::text').get(default='').replace('pcm',
                                                                                                            '').strip()
        item['Price PW'] = selector.css('article div:contains("pw")::text').get(default='').replace('pw',
                                                                                                    '').strip()
        item['Property Type'] = get_value_by_heading(selector, 'PROPERTY TYPE') or json_data.get('propertySubType',
                                                                                                 '')
        item['Bedrooms'] = get_value_by_heading(selector, 'BEDROOMS') or str(json_data.get('bedrooms', ''))
        item['Bathrooms'] = get_value_by_heading(selector, 'BATHROOMS') or str(json_data.get('bathrooms', ''))
        item['Available Date'] = get_value_by_heading(selector, 'Let available date:', letting_details=True)
        item['Furnish Type'] = get_value_by_heading(selector, 'Furnish type:', letting_details=True)
        item['image_urls'] = images_urls
        item['Floor Plan image url'] = floor_plan
        item.update(images)

        data.append(item)

        return data

    except Exception as e:
        print(f"Error processing URL {url}: {e}")
        data = []


if __name__ == "__main__":
    main()
