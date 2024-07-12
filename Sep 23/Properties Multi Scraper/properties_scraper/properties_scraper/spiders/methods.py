import os
from datetime import datetime

from openpyxl.reader.excel import load_workbook  # Load  worksheet
from openpyxl.utils import get_column_letter  # Font Color
from openpyxl.workbook import Workbook  # Worksheet related Tasks


def comparison_excel(data):
    output_dir = 'output/'
    os.makedirs(output_dir, exist_ok=True)
    file_name = f'{output_dir}Data.xlsx'

    try:
        if os.path.isfile(file_name):
            wb = load_workbook(file_name)
        else:
            wb = Workbook()

        if "Sheet" in wb.sheetnames:
            sheet = wb["Sheet"]
            wb.remove(sheet)

        headers = ['Address', 'street number', 'Type', 'Rooms', 'Other', 'Size (m2)', 'Agency url',
                   'Agency name']
        fulldata_headers = ['Id', 'Address', 'street number', 'housing company name', 'Zip code', 'Stairway',
                            'Apartment nr', 'Type', 'Rooms', 'Kitchen', 'Bathroom', 'Sauna', 'Other', 'Size (m2)',
                            'Additional size (m2)', 'Floor', 'Shares', 'Nr of shares', 'Agency url', 'Agency name',
                            'Owned', 'Claimed', 'Status', 'Messages', 'Hero images', 'Description', 'building image']
        addresses_headers = ['Address', 'street number']

        fulldata_sheet = get_work_sheet(sheet_name='FULLDATA', wb=wb, headers=fulldata_headers)
        scrapedata_sheet = get_work_sheet(sheet_name='SCRAPEDATA', wb=wb, headers=headers)
        addresses_sheet = get_work_sheet(sheet_name='ADDRESSES', wb=wb, headers=addresses_headers)
        storage_sheet = get_work_sheet(sheet_name='STORAGE', wb=wb, headers=headers + ['timestamp'])

        timestamp_format = "%Y-%m-%d %H:%M:%S"

        # ********* step 1  ***************

        # Remove all rows in FULLDATA worksheet where there is any content in column S (agency_url)
        remove_fulldata_agencyurl_columns(fulldata_sheet)

        # ********* Step 2: Remove All Rows in SCRAPEDATA ***************

        if 'SCRAPEDATA' in wb.sheetnames:
            scrapedata_sheet = wb['SCRAPEDATA']
            scrapedata_sheet.delete_rows(2, scrapedata_sheet.max_row)
            print('All Rows Deleted from SCRAPEDATA Sheet Successfully')

        # *********  Add Scrape Results from Spiders into SCRAPEDATA worksheet ***************
        for item in data:
            if not item.get('Agency url', ''):
                continue  # Skip this item if 'Agency url' is empty

            # Convert 'street number' to an integer if it exists
            street_no = item.get('street number', '')
            if street_no.isdigit():
                item['street number'] = int(street_no)

            scrapedata_sheet.append([item.get(col) for col in headers])

        # ********* Step 3  ***************

        # Check if the headers are in fulldata_headers, and if not, use scrapedata_headers
        if "Address" in fulldata_headers and "street number" in fulldata_headers:
            column_address = fulldata_headers.index("Address") + 1  # Adding 1 because column indexes are 1-based
            column_street_number = fulldata_headers.index("street number") + 1
        else:
            column_address = headers.index("Address") + 1  # Use scrapedata_headers
            column_street_number = headers.index("street number") + 1

        # Iterates row by row from Scrapedata worksheet for next processing.
        for scrapeddata_row in scrapedata_sheet.iter_rows(min_row=2, max_row=scrapedata_sheet.max_row):

            scrapeddata_values_appendable = [cell.value for cell in scrapeddata_row]
            scrapeddata_values = [str(cell.value).lower().strip() for cell in scrapeddata_row]

            # Start Checking if columns Address(B) and Street No (C) in SCRAPEDATA match any row in FULLDATA
            matching_row = None

            # check for fulldata sheet is not empty
            if fulldata_headers and column_address <= len(fulldata_headers) and column_street_number <= len(
                    fulldata_headers):
                for fulldata_row in fulldata_sheet.iter_rows(min_row=2, max_row=fulldata_sheet.max_row,
                                                             min_col=column_address, max_col=column_street_number):
                    fulldata_values = [str(cell.value).lower().strip() for cell in fulldata_row]

                    # If scrapedata rows matched in fulldata sheet address and street no
                    if fulldata_values == scrapeddata_values[0:2]:  # comparison address and street no, 0 to 2
                        matching_row = fulldata_row[0].row

                        # Shift existing rows down to make space for the new row
                        fulldata_sheet.insert_rows(matching_row, 1)

                        # Insert scraped data_values as a new row at the matching position
                        for header, value in zip(headers, scrapeddata_values_appendable):
                            if header in fulldata_headers:
                                col_index = fulldata_headers.index(header) + 1  # Find the column index for the header
                                fulldata_sheet.cell(row=matching_row, column=col_index, value=value)
                        break

            # Check if it's still None (no match in FULL DATA Sheet)
            if matching_row is None:

                # Initialize a flag to track whether the row is matched in the addresses worksheet
                matched_in_addresses = False

                # column_address ==1, column_street_number ==2
                for addresses_row in addresses_sheet.iter_rows(min_row=2, max_row=addresses_sheet.max_row, min_col=1,
                                                               max_col=2):

                    addresses_values = [str(cell.value).lower().strip() for cell in
                                        addresses_row]

                    if addresses_values == scrapeddata_values[0:2]:

                        data_to_append = []

                        for header in fulldata_headers:
                            if header in headers:
                                col_index = headers.index(header) + 1
                                if col_index <= len(scrapeddata_values):
                                    # value = scrapeddata_values[col_index - 1]
                                    value = scrapeddata_values_appendable[col_index - 1]
                                    if value == 'none' or value == '':
                                        data_to_append.append(None)  # Replace empty string with None
                                    else:
                                        data_to_append.append(value)
                                else:
                                    data_to_append.append(
                                        None)  # Append None if no matching value found in scrapeddata_values
                            else:
                                data_to_append.append(None)  # Append None for headers that don't match

                        # Append the data to FULLDATA

                        fulldata_sheet.append(data_to_append)
                        matched_in_addresses = True
                        break

                if not matched_in_addresses:
                    print(f'Scrapedata row {scrapeddata_values} not matched in Addresses worksheet and Ignored.')

        # Insert rows into Storage those are added in Full data and remove the duplicates
        storage_existing_data = {}
        for storage_row in storage_sheet.iter_rows(min_row=2, max_row=storage_sheet.max_row, min_col=1,
                                                   max_col=8):  # column_address ==1, Agency name ==7
            key = tuple([cell.value for cell in storage_row])
            storage_existing_data[key] = [cell.value for cell in storage_row]

        # Create a mapping between column names in fulldata_headers and headers
        column_mapping = {}
        for header in fulldata_headers:
            if header in headers:
                column_mapping[header] = headers.index(header)

        # Iterate through rows in FULLDATA
        for row in fulldata_sheet.iter_rows(min_row=2, max_row=fulldata_sheet.max_row,
                                            min_col=1, max_col=len(fulldata_headers)):

            # Skip those rows from FullData , it's already existing
            if not row[19].value:
                continue

            fulldata_head = [fulldata_headers[cell.column - 1] for cell in row]
            fulldata_values = [cell.value for cell, header in zip(row, fulldata_head) if header in column_mapping]

            # Check if the row already exists in STORAGE
            if tuple(fulldata_values) not in storage_existing_data:
                timestamp = datetime.now().strftime(timestamp_format)
                storage_sheet.append(fulldata_values + [timestamp])
                storage_existing_data[tuple(fulldata_values)] = fulldata_values

        # Set column widths for All sheet
        column_widths = {
            'Address': 25,
            'street number': 15,
            'Type': 20,
            'Rooms': 10,
            'Other': 10,
            'Size (m2)': 15,
            'Agency url': 50,
            'Agency name': 15,
            'timestamp': 20,
        }

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for header_name, width in column_widths.items():
                for column_letter, cell in enumerate(sheet[1], start=1):
                    if cell.value == header_name:
                        if width is not None:
                            sheet.column_dimensions[get_column_letter(column_letter)].width = width

        wb.save(file_name)
        print(f" Successfully created output File :{file_name}")

    except Exception as e:
        print(f"An error occurred while saving the data: {str(e)}")

    return file_name


def get_work_sheet(sheet_name, wb, headers):
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(sheet_name)
        sheet.append(headers)

    return sheet


def remove_fulldata_agencyurl_columns(fulldata_sheet):
    fulldata_headers = [cell.value for cell in fulldata_sheet[1]]
    if fulldata_headers and 'Agency url' in fulldata_headers:

        # Remove all rows in FULLDATA worksheet where there is any content in column S (agency_url)
        column_agency_url = fulldata_headers.index("Agency url") + 1  # Adding 1 because column indexes are 1-based
        fulldata_rows_to_delete = []

        for row in fulldata_sheet.iter_rows(min_row=2, max_row=fulldata_sheet.max_row, min_col=column_agency_url,
                                            max_col=column_agency_url):
            if row[0].value:
                fulldata_rows_to_delete.append(row[0].row)

        for row_number in reversed(fulldata_rows_to_delete):
            fulldata_sheet.delete_rows(idx=row_number)

    return print('All Agency Url rows removed')
